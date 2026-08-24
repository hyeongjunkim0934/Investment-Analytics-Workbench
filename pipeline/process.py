#!/usr/bin/env python3
"""Investment Analytics Workbench — data pipeline.

Reads vendor Excel exports from the (private) data repository checkout and
emits the JSON datasets consumed by the static dashboard.

공개 범위: 원본 엑셀 파일과 전체 시리즈는 비공개 저장소에만 존재한다.
이 스크립트가 게시하는 것은 패널 정의(OVERVIEW_CARDS, CURVES 등)에 선별된
시리즈의 값(최근 구간 일별, 과거 구간 주별 축약)과 파생 지표(스프레드 등)이며,
카탈로그는 값 없이 메타데이터만 게시한다. 게시된 값은 공개 페이지에서 누구나
접근 가능하므로, 공개해도 되는 시리즈만 패널에 넣어야 한다.

Usage:
    python pipeline/process.py --data-dir source-data --out _site/data
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

import alloc
import bm
import breadth
import estimate
import hedge
import panel
import port
import risk
# 재수출(re-export): 예전 process.epoch_seconds 를 쓰던 호출부를 그대로 두면서
# 정의는 common 한 곳만 남긴다. tests/test_formulas.py 가 동일성을 단정한다.
from common import epoch_seconds

# --------------------------------------------------------------------------
# series store
# --------------------------------------------------------------------------

SERIES: dict[str, dict] = {}   # key -> {source, category, name, s: pd.Series}
WARNINGS: list[str] = []
MERGED_KEYS: set[str] = set()


def warn(msg: str) -> None:
    WARNINGS.append(msg)
    print(f"[warn] {msg}", file=sys.stderr)


def add_series(key: str, source: str, category: str, name: str, pairs) -> None:
    if not pairs:
        return
    idx = pd.DatetimeIndex([p[0] for p in pairs])
    s = pd.Series([p[1] for p in pairs], index=idx, dtype="float64")
    s = s.sort_index()
    s = s[~s.index.duplicated(keep="last")].dropna()
    if s.empty:
        return
    if key in SERIES:
        # 같은 시리즈가 여러 파일에 존재하면 병합하되, 마지막 관측일이
        # 늦은(더 최신인) 쪽이 겹치는 날짜에서 우선한다.
        old = SERIES[key]["s"]
        loser, winner = (old, s) if s.index[-1] >= old.index[-1] else (s, old)
        merged = pd.concat([loser, winner])
        merged = merged[~merged.index.duplicated(keep="last")].sort_index()
        SERIES[key]["s"] = merged
        MERGED_KEYS.add(key)
        return
    SERIES[key] = {"source": source, "category": category, "name": name, "s": s}


def get(key: str) -> pd.Series | None:
    entry = SERIES.get(key)
    if entry is None:
        warn(f"series not found: {key}")
        return None
    return entry["s"]


# --------------------------------------------------------------------------
# cell helpers
# --------------------------------------------------------------------------

def to_num(v):
    """Vendor cell -> float or None. '#N/A N/A', '#NAME?', '' -> None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    if isinstance(v, str):
        t = v.strip().replace(",", "")
        if not t or t.startswith("#"):
            return None
        try:
            return float(t)
        except ValueError:
            return None
    return None


def is_dt(v) -> bool:
    return isinstance(v, datetime)


def cell_label(v) -> str | None:
    return v.strip() if isinstance(v, str) else None


# --------------------------------------------------------------------------
# parsers
# --------------------------------------------------------------------------

CAT_LABELS = {"카테고리", "Category", "CATEGORY"}
NOTATION_LABELS = {"Notation", "NOTATION", "노테이션"}


def parse_wide(path: Path, source_tag: str) -> int:
    """Parse a wide export (Bloomberg / Infomax style).

    Layout: a header block somewhere in the first rows containing a category
    row (col A == 카테고리/Category) and a notation row (col A == Notation),
    followed by data rows whose col A is a datetime. Everything else is
    ignored, so extra vendor rows (Ticker, field, item labels, broken formula
    rows) are harmless.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    added = 0
    try:
        for ws in wb.worksheets:
            cats = notations = None
            cols: list[tuple[int, str, str]] = []   # (col_idx, category, name)
            data: dict[int, list] = {}
            dates: list[datetime] = []
            started = False
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                a = row[0]
                if not started:
                    label = cell_label(a)
                    if label in CAT_LABELS:
                        cats = row
                    elif label in NOTATION_LABELS:
                        notations = row
                    elif is_dt(a) and notations is not None:
                        # header block complete -> build column map, start data
                        seen: set[str] = set()
                        for j, name in enumerate(notations):
                            if j == 0:
                                continue
                            nm = cell_label(name)
                            if not nm:
                                continue
                            if nm in seen:
                                warn(f"{path.name}/{ws.title}: duplicate column '{nm}' skipped")
                                continue
                            seen.add(nm)
                            cat = ""
                            if cats is not None and j < len(cats):
                                cat = cell_label(cats[j]) or ""
                            cols.append((j, cat, nm))
                            data[j] = []
                        started = True
                    else:
                        continue
                if started and is_dt(a):
                    dates.append(a)
                    n = len(row)
                    for j, _, _ in cols:
                        data[j].append(to_num(row[j]) if j < n else None)
            for j, cat, nm in cols:
                key = f"{source_tag}:{nm}"
                pairs = [(d, v) for d, v in zip(dates, data[j])]
                before = len(SERIES)
                add_series(key, source_tag, cat, nm, pairs)
                added += len(SERIES) - before
    finally:
        wb.close()
    return added


def parse_index_export(path: Path) -> tuple[str | None, list]:
    """Parse a vendor index export (name in A1, '일자/종가/...' table below)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = None
    pairs = []
    try:
        ws = wb.worksheets[0]
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            a = row[0]
            if name is None:
                lbl = cell_label(a)
                if lbl:
                    name = lbl.split()[0]
                continue
            if not header_seen:
                if cell_label(a) == "일자":
                    header_seen = True
                continue
            if is_dt(a):
                close = to_num(row[1]) if len(row) > 1 else None
                if close is not None:
                    pairs.append((a, close))
    finally:
        wb.close()
    return name, pairs


def load_data_dir(data_dir: Path) -> list[dict]:
    """Discover and parse every workbook in the data repo checkout."""
    files_report = []
    all_paths = [p for p in data_dir.rglob("*")
                 if p.is_file() and not p.name.startswith("~$")]
    # 대소문자 무관 확장자 매칭 (DATA_BB.XLSX 등도 인식)
    xlsx = sorted(p for p in all_paths if p.suffix.lower() in (".xlsx", ".xlsm"))
    for p in all_paths:
        if p.suffix.lower() == ".xls":
            warn(f"legacy .xls 형식은 지원되지 않아 건너뜀: {p.name} — .xlsx로 다시 저장해 업로드하세요")
            files_report.append({"file": p.name, "kind": "skipped-xls", "series": 0})
    if not xlsx:
        print(f"error: no .xlsx files under {data_dir}", file=sys.stderr)
        sys.exit(1)

    index_files = []   # (name, pairs, max_date, filename)
    for p in xlsx:
        low = p.name.lower()
        if low.startswith("data_bb"):
            n = parse_wide(p, "bb")
            files_report.append({"file": p.name, "kind": "bloomberg-wide", "series": n})
        elif low.startswith("data_info"):
            n = parse_wide(p, "info")
            files_report.append({"file": p.name, "kind": "infomax-wide", "series": n})
        elif bm.is_bm(p):
            # 자산군 전략 벤치마크 — 자산배분 CMA 의 원천. 날짜 박힌 이름이어도
            # 접두사 `bm` 만 지키면 된다 (규약은 pipeline/bm.py 모듈 docstring).
            cols = bm.parse(p, warn)
            for name, grp, pairs in cols:
                # 그룹 행이 비어 있으면 grp="" — 키에 선행 공백이 남지 않게 잇는다
                label = " ".join(x for x in (grp, name) if x)
                add_series(f"bm:{label}", "user-bm", "BM",
                           f"{label} 전략 벤치마크", pairs)
            # 관측이 실제로 있는 열만 센다 — 날짜 행을 못 읽은 파일이 "10개 파싱"
            # 으로 보고되면 meta 와 실제 카탈로그가 어긋난다(재점검 발견).
            files_report.append({"file": p.name, "kind": "benchmark",
                                 "series": sum(1 for _n, _g, pr in cols if pr)})
        elif breadth.is_stock_report(p):
            # 미국 증시 데일리 리포트 — **하루치 스냅샷**이라 성격이 다르다.
            # 파일 하나 = 관측 하루이고, 이력은 날짜별 파일이 쌓여야 생긴다.
            # 판정은 파일명이 아니라 시트 이름으로 한다(리포트는 날짜가 박힌 이름으로
            # 배포되므로 접두사 규약을 매일 지키게 하면 언젠가 반드시 잊는다).
            rep = breadth.parse(p, warn)
            if rep:
                ts = pd.Timestamp(rep["date"])
                for k, v in rep["values"].items():
                    label, unit, _ = breadth.METRICS.get(k, (k, "", None))
                    add_series(f"us:{k}", "kiwoom", "US Breadth",
                               f"{label} ({unit})" if unit else label, [(ts, v)])
                files_report.append({"file": p.name, "kind": "us-breadth",
                                     "series": len(rep["values"]), "asof": rep["date"]})
            else:
                files_report.append({"file": p.name, "kind": "skipped", "series": 0})
        else:
            name, pairs = parse_index_export(p)
            if name and pairs:
                mx = max(d for d, _ in pairs)
                index_files.append((name, pairs, mx, p.name))
                files_report.append({"file": p.name, "kind": f"index:{name}",
                                     "series": 1, "rows": len(pairs)})
            else:
                warn(f"unrecognized workbook skipped: {p.name}")
                files_report.append({"file": p.name, "kind": "skipped", "series": 0})

    # merge index exports per index name; the file with the later end date
    # wins on overlapping dates
    by_name: dict[str, list] = {}
    for name, pairs, mx, fname in sorted(index_files, key=lambda x: x[2]):
        by_name.setdefault(name, []).append(pairs)
    for name, chunks in by_name.items():
        merged: dict = {}
        for pairs in chunks:
            for d, v in pairs:
                merged[d] = v
        add_series(f"idx:{name}", "idx", "Index", name, list(merged.items()))

    if MERGED_KEYS:
        warn(f"{len(MERGED_KEYS)}개 시리즈가 여러 파일에 중복되어 병합됨(최신 파일 우선)")
    return files_report


# --------------------------------------------------------------------------
# analytics helpers
# --------------------------------------------------------------------------

def asof(s: pd.Series, ts) -> tuple[pd.Timestamp | None, float | None]:
    sub = s.loc[:ts]
    if len(sub) == 0:
        return None, None
    return sub.index[-1], float(sub.iloc[-1])


def pack(s: pd.Series | None, round_to: int = 4,
         daily_years: int = 5) -> dict | None:
    """Series -> {t:[unix sec], v:[...]}.

    최근 daily_years년은 일별 그대로, 그 이전 구간은 주별(금요일 라벨)로
    축약한다 — 공개 페이로드를 줄이고 문서의 축약 정책과 일치시킨다.
    """
    if s is None or len(s) == 0:
        return None
    s = s.dropna()
    if len(s) == 0:
        return None
    cutoff = s.index[-1] - pd.DateOffset(years=int(daily_years))
    old = s[s.index < cutoff]
    if len(old):
        old = old.resample("W-FRI").last().dropna()
        # 마지막 부분 주의 금요일 라벨이 경계를 넘어 일별 구간과 겹치거나
        # 역순 타임스탬프를 만들지 않도록 잘라낸다.
        old = old[old.index < cutoff]
        s = pd.concat([old, s[s.index >= cutoff]])
    return {"t": epoch_seconds(s.index),
            "v": [round(float(v), round_to) for v in s.values]}


def spark(s: pd.Series | None, points: int = 30) -> dict | None:
    if s is None or len(s) == 0:
        return None
    w = s.resample("W-FRI").last().dropna().tail(points)
    if len(w) < 2:
        w = s.tail(points)
    return {"t": epoch_seconds(w.index),
            "v": [round(float(v), 4) for v in w.values]}


def series_group(defs: list[tuple[str, str]], **pack_kw) -> list[dict]:
    """[(key, label), ...] -> [{key,label,t,v}, ...], skipping missing."""
    out = []
    for key, label in defs:
        p = pack(get(key), **pack_kw)
        if p:
            out.append({"key": key, "label": label, **p})
    return out


HORIZONS = [("d1", "1일"), ("w1", "1주"), ("m1", "1개월"),
            ("m3", "3개월"), ("ytd", "YTD"), ("y1", "1년")]


def changes(s: pd.Series, kind: str) -> dict:
    """kind: 'rate' -> bp diffs, 'price' -> % change, 'level' -> point diffs."""
    last_ts = s.index[-1]
    last_v = float(s.iloc[-1])
    targets = {
        "d1": s.index[-2] if len(s) > 1 else None,
        "w1": last_ts - timedelta(days=7),
        "m1": last_ts - pd.DateOffset(months=1),
        "m3": last_ts - pd.DateOffset(months=3),
        "ytd": pd.Timestamp(year=last_ts.year - 1, month=12, day=31),
        "y1": last_ts - pd.DateOffset(years=1),
    }
    out = {}
    for h, tgt in targets.items():
        if tgt is None:
            out[h] = None
            continue
        _, v0 = asof(s, tgt)
        if v0 is None:
            out[h] = None
        elif kind == "rate":
            out[h] = round((last_v - v0) * 100, 1)          # bp
        elif kind == "price":
            out[h] = round((last_v / v0 - 1) * 100, 2) if v0 else None   # %
        else:
            out[h] = round(last_v - v0, 2)                   # points
    return out


# --------------------------------------------------------------------------
# panel builders
# --------------------------------------------------------------------------

# 개요 구역 — 2026-08-13 사용자 지시로 개요가 **시장 화면 전체의 입구**가 되었다
# (금리·IRS·크레딧·FX·물가·ACWI·매크로가 상단 탭에서 내려와 여기로 들어왔다).
# 순서는 사용자가 정한 것이다: 주식 → 금리 → 환율 → 기타.
# `sections` 는 그 구역에서 들어갈 수 있는 화면들이며, 화면 자체는 그대로 살아 있다
# (해시 라우팅·렌더러 불변 — 상단 탭에서만 내려왔다).
OVERVIEW_GROUPS = [
    ("equity", "주식",  ["acwi"]),
    ("rate",   "금리",  ["rates", "irs"]),
    ("fx",     "환율",  ["fx"]),
    ("other",  "기타",  ["credit", "inflation", "macro"]),
]

# key, label, kind, decimals, unit, group, link
#
# `link` = 그 카드를 누르면 열리는 화면(사용자 지시: "겹치는 지표, 예를들어 개요의 ACWI
# 카드를 클릭하면 탭 중에 ACWI 눌러서 나오는 화면"). **전용 화면이 없는 카드는 링크를
# 비운다** — 없는 링크를 지어내지 않는다. 다만 2026-08-24 사용자 지시("다른 카드들도
# 전부 다 그렇게")로 무링크 카드는 이제 **상세 오버레이**가 열린다 — build_overview 가
# hist 를 실어 주고 app.js openOvDetail 이 그린다(무동작 카드는 여전히 금지).
# 구역당 5장(2026-08-23 사용자 지시), 구역 안 순서는 국가 중요도(한국 > 미국 > 유럽 >
# 일본 > 중국). 주식 구역의 유럽·일본·중국 **주가지수는 원본 데이터에 없어서** 전세계
# (ACWI)와 변동성 2종으로 채웠다 — 지수 컬럼이 익스포트에 추가되면 여기서 교체한다
# (HANDOVER §6 데이터 요청).
OVERVIEW_CARDS = [
    ("bb:한국_KOSPI_PR",          "KOSPI",            "price", 1, "",   "equity", ""),
    ("bb:미국_S&P500_PR",         "S&P 500",          "price", 1, "",   "equity", ""),
    ("idx:ACWI",                 "MSCI ACWI",        "price", 1, "",   "equity", "acwi"),
    ("info:VKOSPI",              "VKOSPI",           "level", 1, "",   "equity", ""),
    ("bb:미국_변동성지수_VIX",    "VIX",              "level", 1, "",   "equity", ""),
    ("info:한국_3y",             "국고 3년",          "rate",  3, "%",  "rate",   "rates"),
    ("info:한국_10y",            "국고 10년",         "rate",  3, "%",  "rate",   "rates"),
    ("info:UST10y",              "미국채 10년",       "rate",  3, "%",  "rate",   "rates"),
    ("info:한국_기준금리",        "한국 기준금리",     "rate",  2, "%",  "rate",   "rates"),
    ("bb:미국_기준금리",          "미국 기준금리",     "rate",  2, "%",  "rate",   "rates"),
    ("info:USDKRW",              "달러/원",           "price", 1, "",   "fx",     "fx"),
    ("bb:달러지수",               "달러지수(DXY)",     "price", 1, "",   "fx",     "fx"),
    ("info:EURKRW",              "유로/원",           "price", 1, "",   "fx",     "fx"),
    ("info:KRWJPY",              "원/100엔",          "price", 1, "",   "fx",     "fx"),
    ("info:USDCNY",              "달러/위안",         "price", 2, "",   "fx",     "fx"),
    ("bb:한국_CDS_5년물",         "한국 CDS 5년",     "level", 1, "bp", "other",  "credit"),
    ("bb:미국_하이일드_스프레드", "미 HY 스프레드",    "rate",  2, "%p", "other",  "credit"),
    ("bb:미국_투자등급_스프레드", "미 IG 스프레드",    "rate",  2, "%p", "other",  "credit"),
    ("bb:WTI유가",                "WTI 유가",         "price", 1, "$",  "other",  ""),
    ("bb:원자재지수",             "원자재(GSCI)",      "price", 1, "",   "other",  ""),
]


def build_overview() -> dict:
    cards = []
    for key, label, kind, dec, unit, group, link in OVERVIEW_CARDS:
        s = get(key)
        if s is None:
            continue
        card = {
            "key": key, "label": label, "kind": kind, "unit": unit,
            "group": group, "link": link,
            "value": round(float(s.iloc[-1]), dec),
            "date": s.index[-1].strftime("%Y-%m-%d"),
            "chg": changes(s, kind),
            "spark": spark(s),
        }
        # 전용 화면이 없는 카드는 오버레이 상세가 열린다(2026-08-24 사용자 지시
        # "전부 다 그렇게") — 그 차트에 쓸 이력을 함께 싣는다(최근 1년 일별 + 이전 주별)
        if not link:
            hist = pack(s, daily_years=1)
            if hist:
                card["hist"] = hist
        cards.append(card)
    # 구역은 **카드가 하나라도 살아남은 것만** 내보낸다 — 시리즈가 빠져 텅 빈 구역
    # 제목이 화면에 남으면 "여기 뭔가 있어야 하는데 안 나온다"로 읽힌다.
    live = {c["group"] for c in cards}
    groups = [{"key": g, "label": lb, "sections": secs}
              for g, lb, secs in OVERVIEW_GROUPS if g in live]
    return {"cards": cards, "groups": groups}


CURVES = {
    "KR": {"label": "한국 국고", "src": "info", "prefix": "한국_",
           "tenors": ["3m", "6m", "9m", "1y", "1.5y", "2y", "3y", "4y", "5y",
                      "7y", "10y", "15y", "20y", "30y"]},
    "US": {"label": "미국 국채", "src": "info", "prefix": "UST",
           "tenors": ["3m", "6m", "1y", "2y", "3y", "5y", "7y", "10y",
                      "20y", "30y"]},
    "JP": {"label": "일본 국채", "src": "info", "prefix": "JPY",
           "tenors": ["1y", "2y", "3y", "4y", "5y", "7y", "10y", "15y",
                      "20y", "30y", "40y"]},
    "DE": {"label": "독일 국채", "src": "info", "prefix": "GER",
           "tenors": ["1y", "2y", "3y", "5y", "7y", "10y", "20y", "30y"]},
    "AU": {"label": "호주 국채", "src": "info", "prefix": "AUD",
           "tenors": ["1y", "2y", "3y", "5y", "7y", "10y", "20y", "30y"]},
}

def curve_series_key(code: str, tenor: str) -> str:
    cfg = CURVES[code]
    return f"{cfg['src']}:{cfg['prefix']}{tenor}"


def tenor_years(t: str) -> float:
    t = t.lower()
    if t.endswith("m"):
        return float(t[:-1]) / 12
    return float(t[:-1])


def build_rates() -> dict:
    curves = {}
    for code, cfg in CURVES.items():
        tenors, xs, series_list = [], [], []
        for t in cfg["tenors"]:
            s = get(curve_series_key(code, t))
            if s is not None:
                tenors.append(t)
                xs.append(round(tenor_years(t), 4))
                series_list.append(s)
        if not series_list:
            continue
        last_ts = max(s.index[-1] for s in series_list)
        snap_defs = [("최근", last_ts), ("1개월 전", last_ts - pd.DateOffset(months=1)),
                     ("1년 전", last_ts - pd.DateOffset(years=1))]
        snaps = []
        for lbl, tgt in snap_defs:
            vals, dt = [], None
            for s in series_list:
                d, v = asof(s, tgt)
                vals.append(None if v is None else round(v, 3))
                if d is not None and (dt is None or d > dt):
                    dt = d
            snaps.append({"label": lbl,
                          "date": dt.strftime("%Y-%m-%d") if dt is not None else None,
                          "v": vals})
        curves[code] = {"label": cfg["label"], "tenors": tenors, "x": xs,
                        "snaps": snaps}

    ts10 = series_group([
        ("info:한국_10y", "한국"), ("info:UST10y", "미국"),
        ("info:JPY10y", "일본"), ("info:GER10y", "독일"),
    ])
    policy = series_group([
        ("info:한국_기준금리", "한국"), ("bb:미국_기준금리", "미국"),
        ("bb:유로_기준금리", "유로"), ("bb:일본_기준금리", "일본"),
    ])

    spreads = []
    for label, a_key, b_key in [
        ("한국 10년−3년", "info:한국_10y", "info:한국_3y"),
        ("미국 10년−2년", "info:UST10y", "info:UST2y"),
        ("한미 10년 금리차", "info:한국_10y", "info:UST10y"),
    ]:
        a, b = get(a_key), get(b_key)
        if a is None or b is None:
            continue
        sp = ((a - b) * 100).dropna()     # bp
        p = pack(sp, round_to=1)
        if p:
            spreads.append({"label": label, **p})
    return {"curves": curves, "ts10": ts10, "policy": policy, "spreads": spreads}


IRS_TENORS = ["1y3m", "1y1y", "2y1y", "3y1y", "4y1y", "5y1y", "5y5y", "5y10y"]
IRS_COUNTRIES = [("미국", "미국"), ("한국", "한국"), ("일본", "일본"), ("유로", "유로")]


def build_irs() -> dict:
    fwd = {}
    for code, label in IRS_COUNTRIES:
        tenors, series_list = [], []
        for t in IRS_TENORS:
            s = get(f"bb:{code}_IRS_{t}")
            if s is not None:
                tenors.append(t)
                series_list.append(s)
        if not series_list:
            continue
        last_ts = max(s.index[-1] for s in series_list)
        snaps = []
        for lbl, tgt in [("최근", last_ts),
                         ("1개월 전", last_ts - pd.DateOffset(months=1)),
                         ("1년 전", last_ts - pd.DateOffset(years=1))]:
            vals, dt = [], None
            for s in series_list:
                d, v = asof(s, tgt)
                vals.append(None if v is None else round(v, 3))
                if d is not None and (dt is None or d > dt):
                    dt = d
            snaps.append({"label": lbl,
                          "date": dt.strftime("%Y-%m-%d") if dt is not None else None,
                          "v": vals})
        fwd[code] = {"label": label, "tenors": tenors, "snaps": snaps}

    ts = series_group([
        ("bb:미국_IRS_1y1y", "미국 1y1y"), ("bb:미국_IRS_5y5y", "미국 5y5y"),
        ("bb:한국_IRS_1y1y", "한국 1y1y"), ("bb:한국_IRS_5y5y", "한국 5y5y"),
    ])
    return {"fwd": fwd, "ts": ts}


KR_CREDIT_3Y = [
    ("info:Public_AAA_3y", "공사채 AAA"),
    ("info:KDB_AAA_3y", "산금채 AAA"),
    ("info:Bank_AAA_3y", "은행채 AAA"),
    ("info:Corp_AAA_3y", "회사채 AAA"),
    ("info:Corp_AA_zero_3y", "회사채 AA0"),
    ("info:Corp_A_plus_3y", "회사채 A+"),
    ("info:Card_AA_plus_3y", "카드채 AA+"),
    ("info:capital_A_plus_3y", "캐피탈채 A+"),
]


def build_credit() -> dict:
    base = get("info:한국_3y")
    kr = []
    if base is not None:
        for key, label in KR_CREDIT_3Y:
            s = get(key)
            if s is None:
                continue
            sp = ((s - base) * 100).dropna()   # bp over KTB 3y
            p = pack(sp, round_to=1)
            if p:
                kr.append({"key": key, "label": label, **p})
    us = series_group([
        ("bb:미국_투자등급_스프레드", "미국 IG"),
        ("bb:미국_하이일드_스프레드", "미국 HY"),
    ])
    cds = series_group([
        ("bb:한국_CDS_5년물", "한국"), ("bb:일본_CDS_5년물", "일본"),
        ("bb:중국_CDS_5년물", "중국"), ("bb:독일_CDS_5년물", "독일"),
        ("bb:프랑스_CDS_5년물", "프랑스"),
    ])
    return {"kr": kr, "us": us, "cds": cds}


def build_fx() -> dict:
    # 원/100엔·달러/위안은 개요 환율 카드의 링크 도착지가 이 화면이라 함께 그린다
    # (2026-08-24 — 카드가 가리키는 화면에 그 계열의 차트가 실제로 있어야 한다)
    ts = series_group([
        ("info:USDKRW", "달러/원"), ("info:DXY", "달러지수"),
        ("info:USDJPY", "달러/엔"), ("info:EURKRW", "유로/원"),
        ("info:KRWJPY", "원/100엔"), ("info:USDCNY", "달러/위안"),
    ])
    # 헤지비용(HP 커브)은 여기서 싣지 않는다 — `hedge.py` 가 `cost_curve`·
    # `cost_hist_curve` 로 싣고 #hedge 화면이 그린다. 같은 값을 두 JSON 에 실으면
    # 새 이중 진실이 되고, `renderHedge` 가 `DATA.fx` 를 읽으면 fx.json 하나가
    # 깨질 때 #hedge 까지 함께 빈다. #fx 는 순수 시세 화면이다.
    return {"ts": ts}


def build_inflation() -> dict:
    bei = series_group([
        ("info:KTB_BEI10y", "한국"), ("info:UST_BEI10y", "미국"),
        ("info:JGB_BEI10y", "일본"), ("info:GER_BEI10y", "독일"),
        ("info:AUD_BEI10y", "호주"),
    ])
    tips = series_group([
        ("bb:미국_TIPS_10y", "미국 TIPS 10y"), ("bb:한국_TIPS_10y", "한국 TIPS 10y"),
    ])
    return {"bei": bei, "tips": tips}


def build_acwi() -> dict:
    s = get("idx:ACWI")
    if s is None:
        return {}
    price = pack(s, round_to=3, daily_years=8)
    peak = s.cummax()
    dd = ((s / peak - 1) * 100).round(2)
    drawdown = pack(dd, round_to=2, daily_years=8)

    last_ts = s.index[-1]
    first_ts = s.index[0]
    years = (last_ts - first_ts).days / 365.25
    cagr = ((float(s.iloc[-1]) / float(s.iloc[0])) ** (1 / years) - 1) * 100 if years > 0 else None
    daily_ret = s.pct_change().dropna()
    vol = float(daily_ret.tail(252).std()) * math.sqrt(252) * 100 if len(daily_ret) > 30 else None
    stats = {
        "last": round(float(s.iloc[-1]), 2),
        "date": last_ts.strftime("%Y-%m-%d"),
        "first_date": first_ts.strftime("%Y-%m-%d"),
        "cagr": round(cagr, 2) if cagr is not None else None,
        "vol_1y": round(vol, 2) if vol is not None else None,
        "mdd": round(float(dd.min()), 2),
        "high_52w": round(float(s.loc[last_ts - timedelta(days=365):].max()), 2),
        "low_52w": round(float(s.loc[last_ts - timedelta(days=365):].min()), 2),
        "chg": changes(s, "price"),
    }
    return {"price": price, "drawdown": drawdown, "stats": stats,
            "breadth": build_breadth()}


def build_breadth() -> dict:
    """미국 증시 시장 폭 — 이 화면이 답하지 못하던 질문 하나를 채운다.

    ACWI 는 **지수 하나의 가격**이라 "올랐다/내렸다"만 말할 수 있고, 그 상승이
    전 종목이 함께 오른 것인지 대형주 몇 개가 끈 것인지는 답하지 못한다.
    시장 폭은 그 구분을 준다.

    **관측이 하루뿐일 수 있다** — 원본이 데일리 리포트라 이력은 날짜별 파일이
    쌓여야 생긴다. 그래서 `n` 을 그대로 실어 화면이 스냅샷과 추세를 구분하게 한다.
    없는 이력을 있는 것처럼 보이는 차트를 그리지 않기 위한 필드다.
    """
    rows, ts_out, n_obs = [], {}, 0
    for key, (label, unit, note) in breadth.METRICS.items():
        entry = SERIES.get(f"us:{key}")
        if entry is None:
            continue
        ser = entry["s"].dropna()
        if ser.empty:
            continue
        n_obs = max(n_obs, len(ser))
        rows.append({"key": key, "label": label, "unit": unit, "note": note,
                     "last": round(float(ser.iloc[-1]), 3),
                     "date": ser.index[-1].strftime("%Y-%m-%d"),
                     "n": int(len(ser))})
        if len(ser) >= 2:
            # 점 하나짜리 시계열은 싣지 않는다 — 차트가 그려질 수 없고,
            # 실려 있으면 화면이 "이력이 있다"고 오판한다.
            ts_out[key] = pack(ser, round_to=3, daily_years=50)
    if not rows:
        return {}
    return {"asof": max(r["date"] for r in rows), "n": n_obs,
            "rows": rows, "ts": ts_out,
            "src": "키움증권 미국 증시 데일리 리포트 (집계 지표만 — 종목 단위 미게시)"}


MACRO_DEFS = [
    ("bb:한국_GDP_real_yoy", "한국 실질 GDP YoY", "%"),
    ("bb:미국_GDP_real_yoy", "미국 실질 GDP YoY", "%"),
    ("bb:한국_core_CPI_yoy", "한국 근원 CPI YoY", "%"),
    ("bb:미국_core_PCE_CPI_yoy", "미국 근원 PCE YoY", "%"),
    ("bb:미국_실업률", "미국 실업률", "%"),
    ("bb:미국_NonFarm_Payrolls_mom", "미국 비농업고용 MoM", "천명"),
]


def build_macro() -> dict:
    items = []
    for key, label, unit in MACRO_DEFS:
        s = get(key)
        if s is None:
            continue
        # 벤더 데이터는 발표값이 일별로 이월되므로 월말 기준으로 표본화한다.
        s = s.dropna()
        last_obs_date = s.index[-1]
        s = s.resample("ME").last().dropna()
        # 이미 월별로 희소한 데이터이므로 주별 축약을 적용하지 않는다.
        p = pack(s.tail(60), round_to=3, daily_years=100)
        if p:
            items.append({"key": key, "label": label, "unit": unit,
                          "last": round(float(s.iloc[-1]), 3),
                          "date": last_obs_date.strftime("%Y-%m-%d"), **p})
    return {"items": items}


def build_catalog() -> dict:
    rows = []
    for key, e in sorted(SERIES.items()):
        s = e["s"]
        rows.append({
            "key": key, "source": e["source"], "category": e["category"],
            "name": e["name"],
            "first": s.index[0].strftime("%Y-%m-%d"),
            "last": s.index[-1].strftime("%Y-%m-%d"),
            "n": int(len(s)),
        })
    return {"series": rows}


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()

    files_report = load_data_dir(args.data_dir)
    total = len(SERIES)
    print(f"parsed {total} series from {len(files_report)} files")
    if total < 10:
        print("error: too few series parsed — aborting so a broken upload "
              "does not wipe the dashboard", file=sys.stderr)
        sys.exit(1)

    out = args.out
    out.mkdir(parents=True, exist_ok=True)

    now_utc = datetime.now(timezone.utc)
    kst = now_utc.astimezone(timezone(timedelta(hours=9)))
    last_obs = max(e["s"].index[-1] for e in SERIES.values())

    payloads = {
        "overview.json": build_overview(),
    }

    # 리스크 스코어보드 — 계산 실패가 대시보드 전체 빌드를 막지 않도록 격리
    try:
        risk_payload, events_payload, risk_weekly = risk.build(SERIES, warn)
        payloads["risk.json"] = risk_payload
        # 시장 폭 이벤트를 같은 스트림에 합친다. **risk.py 는 건드리지 않는다** —
        # 규칙의 뜻은 breadth.py 가 알고, 여기서는 합치기만 한다(사용자 제안 2026-08-04).
        # 실패해도 나머지 이벤트는 나가야 하므로 따로 격리한다.
        try:
            b_ev, b_cat = breadth.detect_events(
                SERIES, pd.Timestamp(events_payload["asof"])
                - pd.Timedelta(days=events_payload["lookback_days"]))
            if b_ev:
                events_payload["events"] = sorted(
                    events_payload["events"] + b_ev,
                    key=lambda e: e["date"], reverse=True)[:40]
            if b_cat:
                events_payload["catalog"] = events_payload["catalog"] + b_cat
        except Exception:
            import traceback
            traceback.print_exc()
            warn("breadth: 시장 폭 이벤트 계산 실패 — 나머지 이벤트만 나갑니다")
        # 브리핑 원고 — 반드시 시장 폭 **병합 뒤**에 조립한다(전건이 실려야 한다).
        # 실패하면 브리핑 없이 나가되, 게이트(REQUIRED_KEYS)가 배포 전에 잡는다.
        try:
            events_payload["brief"] = risk.compose_brief(
                events_payload["events"], events_payload["asof"],
                events_payload["lookback_days"])
        except Exception:
            import traceback
            traceback.print_exc()
            warn("brief: 이벤트 브리핑 조립 실패 — 브리핑 카드 없이 배포됩니다")
        payloads["events.json"] = events_payload
        try:
            payloads["panel.json"] = panel.build(SERIES, risk_weekly, warn)
        except Exception:
            import traceback
            traceback.print_exc()
            warn("panel: 관계분석 패널 계산 실패 — 해당 화면 없이 배포됩니다")
    except Exception:
        import traceback
        traceback.print_exc()
        warn("risk: 스코어보드 계산 실패 — 해당 섹션 없이 배포됩니다")

    try:
        payloads["hedge.json"] = hedge.build(SERIES, warn)
    except Exception:
        import traceback
        traceback.print_exc()
        warn("hedge: 환헤지 계산 실패 — 해당 섹션 없이 배포됩니다")

    try:
        payloads["alloc.json"] = alloc.build(SERIES, warn)
    except Exception:
        import traceback
        traceback.print_exc()
        warn("alloc: 자산배분 계산 실패 — 해당 섹션 없이 배포됩니다")

    # CMA(자본시장가정) — BM 파일이 없어도 active:false 로 항상 게시된다
    # (게이트 REQUIRED_KEYS 가 부재를 막으므로, 블록 자체는 무조건 있어야 한다).
    if "alloc.json" in payloads:
        try:
            payloads["alloc.json"]["cma"] = bm.build_cma(SERIES, warn)
        except Exception:
            import traceback
            traceback.print_exc()
            warn("cma: 자본시장가정 계산 실패 — 최적화 카드 없이 배포됩니다")
            payloads["alloc.json"]["cma"] = {"active": False, "reason": "계산 실패 — 빌드 로그 확인"}
        try:
            payloads["alloc.json"]["port"] = port.build(SERIES, warn, args.data_dir)
        except Exception:
            import traceback
            traceback.print_exc()
            warn("port: 포트폴리오 구성 통계 계산 실패 — 해당 패널 없이 배포됩니다")
            payloads["alloc.json"]["port"] = {"active": False, "reason": "계산 실패 — 빌드 로그 확인"}

    # 수익률 추정 화면의 자동 채움 지수(§7.8). 실패해도 화면은 **수기 입력으로 살아 있어야**
    # 하므로 격리하되, 블록 자체는 active:false 로 반드시 게시한다 — 게이트가 부재를 막는다.
    try:
        payloads["estimate.json"] = estimate.build(SERIES, warn)
    except Exception:
        import traceback
        traceback.print_exc()
        warn("estimate: 자동 채움 지수 계산 실패 — 수익률 추정 화면은 수기 입력으로만 동작합니다")
        payloads["estimate.json"] = {
            "active": False, "reason": "계산 실패 — 빌드 로그 확인",
            "indices": [], "unavailable": estimate.UNAVAILABLE,
            "annualize": estimate.ANNUALIZE}

    payloads.update({
        "rates.json": build_rates(),
        "irs.json": build_irs(),
        "credit.json": build_credit(),
        "fx.json": build_fx(),
        "inflation.json": build_inflation(),
        "acwi.json": build_acwi(),
        "macro.json": build_macro(),
        "catalog.json": build_catalog(),
    })
    payloads["meta.json"] = {
        "built_at_utc": now_utc.strftime("%Y-%m-%d %H:%M UTC"),
        "built_at_kst": kst.strftime("%Y-%m-%d %H:%M KST"),
        "last_observation": last_obs.strftime("%Y-%m-%d"),
        "series_count": total,
        "files": files_report,
        "warnings": WARNINGS,
        "note": "선별된 시리즈만 게시됩니다. 최근 구간(기본 5년)은 일별, 그 이전 구간은 주별로 축약됩니다.",
    }

    for fname, obj in payloads.items():
        p = out / fname
        with p.open("w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, separators=(",", ":"),
                      allow_nan=False)
        print(f"wrote {p} ({p.stat().st_size:,} bytes)")

    if WARNINGS:
        print(f"{len(WARNINGS)} warning(s) — see meta.json", file=sys.stderr)


if __name__ == "__main__":
    main()
