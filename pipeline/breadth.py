# -*- coding: utf-8 -*-
"""미국 증시 데일리 리포트 → **집계 지표만** 뽑아 시계열로 쌓는다.

이 파서가 읽는 파일은 다른 입력과 성격이 완전히 다르다:
- `data_bb`·`data_info` 는 **와이드 시계열**(한 파일에 수백 계열 × 25년)이다.
- 이 리포트는 **하루치 스냅샷**이다. 파일 하나 = 관측 하루.
  따라서 이력은 **파일이 쌓여야만** 생긴다(같은 폴더에 날짜별 파일이 여러 개).

**공개 범위 — 이 파서가 뽑지 않는 것이 뽑는 것보다 중요하다.**
리포트에는 티커·회사명·현재가·시가총액·52주 고저 같은 **종목 단위 벤더 데이터**가
7,000종목 규모로 들어 있다. 이 저장소는 공개 저장소이고 파생 JSON 은 GitHub Pages 로
나가므로, **종목 단위는 한 줄도 읽지 않는다.** 여기서 뽑는 것은 전부 시장 전체 또는
지수 단위의 **집계 수**(상승/하락 종목 수, 참여율, 신고가·신저가 종목 수, 평균 변동률)
이며, 이는 개별 종목을 복원할 수 없는 요약 통계다.

**무엇에 쓰는가 — 「지수가 올랐는데 넓은 상승인가, 좁은 상승인가」.**
기존 대시보드에는 이 질문에 답할 수 있는 시리즈가 하나도 없다(가격·금리·스프레드뿐).
`docs/HANDOVER.md` §5.4 가 기록한 대로 현재 위험 지표는 **동행 지표이지 선행 지표가
아닌데**, 시장 폭(breadth)은 몇 안 되는 선행 후보다.

**아직 위험 모형에 넣지 않는다.** `risk.py` 의 요인 점수는 백분위·z 점수와
워크포워드 재적합을 쓰므로 관측 1일로는 계산 자체가 불가능하다. 이력이 쌓이기
전에 넣으면 그것이 곧 자의성이다. 지금은 **게시만** 한다.
"""

from __future__ import annotations

import re

import openpyxl

#: 이 워크북을 알아보는 표식. 파일명이 아니라 **내용**으로 판정한다 —
#: 리포트는 날짜가 박힌 이름으로 배포되므로(`Daily_Stock_Report_26.08.04.xlsx`)
#: 접두사 규약을 매일 지키게 하면 언젠가 반드시 잊는다.
SIGNATURE_SHEET = "Market Overview"

#: 시장 요약 블록의 헤더 — 이 라벨들이 한 행에 같이 있으면 그 다음 행이 값이다.
_MARKET_HEAD = ("상장 보통주", "상승", "하락", "참여율", "신고가권", "신저가권")
#: 지수별 시장 폭 블록의 헤더.
_INDEX_HEAD = ("지수", "종목수", "상승", "하락", "평균(단순)", "평균(시가총액가중)")

#: 게시할 지수 3개 — 대형(S&P500)·기술(NASDAQ)·소형(러셀2000). 셋이 갈라지는 것
#: 자체가 신호라 서로 대체되지 않는다. NYSE only·DOW 는 앞 셋과 크게 겹쳐 뺀다.
INDEX_KEYS = {"S&P500": "sp500", "NASDAQ": "nasdaq", "러셀2000": "r2000"}


def is_stock_report(path) -> bool:
    """워크북이 미국 증시 데일리 리포트인지 — 시트 이름으로만 판정(값을 읽지 않는다)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        return SIGNATURE_SHEET in wb.sheetnames
    finally:
        wb.close()


def _num(v) -> float | None:
    """`"4,596"` · `"74%"` · `2.09` 를 수로. 못 읽으면 None."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        return None if f != f else f
    if not isinstance(v, str):
        return None
    t = v.strip().replace(",", "").replace("%", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _row_has(row, labels) -> bool:
    txt = [str(c).strip() for c in row if isinstance(c, str)]
    return all(any(lb == t for t in txt) for lb in labels)


def _find(rows, labels) -> int:
    """헤더 라벨로 블록을 찾는다 — 행 번호를 박지 않는다.

    리포트 서식은 벤더가 바꿀 수 있고, 행을 박아 두면 서식이 한 줄만 밀려도
    **조용히 다른 블록을 읽는다**(값은 나오는데 뜻이 달라지는 최악의 실패다).
    """
    for i, row in enumerate(rows):
        if _row_has(row, labels):
            return i
    return -1


def parse(path, warn) -> dict:
    """리포트 한 부 → `{"date": "YYYY-MM-DD", "values": {키: 값}}`. 실패하면 빈 dict."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SIGNATURE_SHEET not in wb.sheetnames:
            return {}
        rows = list(wb[SIGNATURE_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()

    # ---- 관측일 — "미국 2026-08-03 종가 기준 …" 에서 뽑는다.
    # 파일명의 날짜(26.08.04)는 **작성일**이라 관측일보다 하루 뒤다. 파일명을 믿으면
    # 모든 관측이 하루씩 밀린다 — 리포트 본문이 스스로 밝힌 날짜를 쓴다.
    date = None
    for row in rows[:8]:
        for c in row:
            if isinstance(c, str):
                m = re.search(r"(\d{4})-(\d{2})-(\d{2})", c)
                if m and "종가" in c:
                    date = m.group(0)
                    break
        if date:
            break
    if date is None:
        warn(f"breadth: 관측일을 찾지 못함 — {getattr(path, 'name', path)}")
        return {}

    out: dict[str, float] = {}

    # ---- 시장 전체 (미국 상장 보통주 전체)
    i = _find(rows, _MARKET_HEAD)
    if i < 0:
        warn(f"breadth: 시장 요약 블록 없음 — {getattr(path, 'name', path)}")
    else:
        head = [str(c).strip() if isinstance(c, str) else "" for c in rows[i]]
        vals = rows[i + 1] if i + 1 < len(rows) else ()
        col = {h: k for k, h in enumerate(head) if h}

        def v(label):
            k = col.get(label)
            return None if k is None or k >= len(vals) else _num(vals[k])

        total, adv, dec = v("상장 보통주"), v("상승"), v("하락")
        nh, nl = v("신고가권"), v("신저가권")
        up5, dn5 = v("급등 +5%"), v("급락 -5%")
        part = v("참여율")

        if adv is not None and dec:
            out["ad_ratio"] = round(adv / dec, 3)
        if part is not None:
            out["participation"] = round(part, 1)
        if nh is not None and nl is not None:
            # 신고가 − 신저가. **이 리포트가 유일하게 답하는 질문**이 여기 있다 —
            # 오른 종목이 훨씬 많은 날에도 이 값이 음수면 상승이 좁다는 뜻이다.
            out["net_new_high"] = round(nh - nl)
            if total:
                out["net_new_high_pct"] = round((nh - nl) / total * 100, 2)
        if up5 is not None and dn5 is not None:
            out["surge_net"] = round(up5 - dn5)
        if total is not None:
            out["universe"] = round(total)

    # ---- 지수별 (상승/하락 비율 + 쏠림)
    j = _find(rows, _INDEX_HEAD)
    if j < 0:
        warn(f"breadth: 지수별 블록 없음 — {getattr(path, 'name', path)}")
    else:
        head = [str(c).strip() if isinstance(c, str) else "" for c in rows[j]]
        col = {h: k for k, h in enumerate(head) if h}
        for row in rows[j + 1: j + 12]:
            if not row or not isinstance(row[0], str):
                continue
            key = INDEX_KEYS.get(row[0].strip())
            if key is None:
                continue

            def cell(label):
                k = col.get(label)
                return None if k is None or k >= len(row) else _num(row[k])

            ad = cell("상승/하락")
            simple, weighted = cell("평균(단순)"), cell("평균(시가총액가중)")
            if ad is not None:
                out[f"ad_{key}"] = round(ad, 2)
            if simple is not None and weighted is not None:
                # 쏠림 = 시총가중 − 단순평균 (%p). 양수면 **대형주가 끌었다**는 뜻,
                # 즉 같은 상승률이라도 폭이 좁다. 리포트는 소수(0.0155)로 주므로 ×100.
                out[f"skew_{key}"] = round((weighted - simple) * 100, 3)

    if not out:
        warn(f"breadth: 뽑아낸 지표가 없음 — {getattr(path, 'name', path)}")
        return {}
    return {"date": date, "values": out}


#: 게시 메타데이터 — 카탈로그·화면이 공유하는 이름과 단위 정본.
METRICS = {
    "ad_ratio":         ("상승/하락 종목수 비율", "배", "1보다 크면 오른 종목이 많다"),
    "participation":    ("참여율", "%", "당일 실제로 거래되어 등락이 잡힌 종목의 비율"),
    "net_new_high":     ("52주 신고가권 − 신저가권", "종목", "음수면 오른 날에도 바닥을 깨는 종목이 더 많다"),
    "net_new_high_pct": ("52주 신고가권 − 신저가권 (전체 대비)", "%", "종목수 차이를 상장 종목수로 나눈 값"),
    "surge_net":        ("급등(+5%) − 급락(−5%)", "종목", "당일 큰 폭 변동의 방향 쏠림"),
    "universe":         ("상장 보통주", "종목", "그날 집계 대상 종목수"),
    "ad_sp500":         ("S&P500 상승/하락", "배", None),
    "ad_nasdaq":        ("NASDAQ 상승/하락", "배", None),
    "ad_r2000":         ("러셀2000 상승/하락", "배", None),
    "skew_sp500":       ("S&P500 쏠림(시총가중 − 단순평균)", "%p", "양수면 대형주가 지수를 끌었다"),
    "skew_nasdaq":      ("NASDAQ 쏠림(시총가중 − 단순평균)", "%p", "양수면 대형주가 지수를 끌었다"),
    "skew_r2000":       ("러셀2000 쏠림(시총가중 − 단순평균)", "%p", "양수면 대형주가 지수를 끌었다"),
}


# ---------------------------------------------------------------------------
# 이벤트 — 「하루 안에서 판정되는 것」만
# ---------------------------------------------------------------------------
# 위험 요인(백분위·z 점수·워크포워드)과 달리 **이력이 필요 없다.** 아래 규칙은 전부
# 같은 날 안에서 서로 다른 두 수를 비교하는 **횡단면 조건**이라, 관측이 하루뿐이어도
# 그날의 판정이 성립한다. 사용자 제안(2026-08-04)이 정확히 이 성질을 짚었다.
#
# **임의 기준을 하나도 쓰지 않는다.** 전부 `>` 비교이거나 "세 지수의 부호가 같은가"
# 같은 만장일치 조건이다. 「신고가가 신저가보다 N개 많으면」 류의 수를 지금 정하면
# 그 N 은 근거 없는 수가 된다 — 이 저장소의 상위 제약(자의성 금지) 위반이다.
# 분포를 세울 만큼 관측이 쌓이면 그때 세기를 나눌 것.

#: 이벤트 카테고리 이름 — `risk.detect_events` 의 카테고리와 겹치지 않게 둔다.
EVENT_CAT = "시장폭"


def detect_events(series_store: dict, win_start) -> tuple[list, list]:
    """`us:*` 시리즈에서 이벤트를 뽑는다 → `(events, catalog_rows)`.

    `win_start` 이후 관측만 본다 — 다른 규칙과 같은 창을 쓴다.
    """
    def ser(key):
        e = series_store.get(f"us:{key}")
        return None if e is None else e["s"].dropna()

    keys = ["ad_ratio", "net_new_high", "universe",
            "ad_sp500", "ad_nasdaq", "ad_r2000",
            "skew_sp500", "skew_nasdaq", "skew_r2000"]
    cols = {k: ser(k) for k in keys}
    if cols["ad_ratio"] is None or cols["net_new_high"] is None:
        return [], []

    dates = [d for d in cols["ad_ratio"].index if d > win_start]
    out = []

    def at(k, d):
        s = cols.get(k)
        if s is None or d not in s.index:
            return None
        return float(s.loc[d])

    for d in dates:
        ad, nnh = at("ad_ratio", d), at("net_new_high", d)
        if ad is None or nnh is None:
            continue
        date_s = d.strftime("%Y-%m-%d")
        univ = at("universe", d)

        # ── ① 시장 폭 괴리 — 방향이 서로 반대다.
        # 오른 종목이 더 많은 날(A/D>1)인데 52주 **바닥**을 깨는 종목이 신고가보다
        # 많다면, 지수가 오른 것과 시장 내부가 튼튼한 것은 다른 이야기다.
        # 두 수 모두 그날 안에서 독립적으로 세어진 것이라 비교에 기준값이 없다.
        if ad > 1 and nnh < 0:
            out.append({
                "date": date_s, "sev": "주의", "cat": EVENT_CAT,
                "title": "미국 증시 — 오른 종목이 많은데 52주 신저가가 더 많음",
                "value": f"상승/하락 {ad:.2f}배 · 신고가−신저가 {nnh:+.0f}종목"
                         + (f" (상장 {univ:,.0f})" if univ else ""),
                "rule": "같은 날 A/D>1 이면서 (52주 신고가권 − 신저가권)<0 — 임계값 없음(부호 비교)",
                "tags": ["breadth", "equity"]})
        elif ad < 1 and nnh > 0:
            out.append({
                "date": date_s, "sev": "정보", "cat": EVENT_CAT,
                "title": "미국 증시 — 내린 종목이 많은데 52주 신고가가 더 많음",
                "value": f"상승/하락 {ad:.2f}배 · 신고가−신저가 {nnh:+.0f}종목",
                "rule": "같은 날 A/D<1 이면서 (52주 신고가권 − 신저가권)>0 — 임계값 없음(부호 비교)",
                "tags": ["breadth", "equity"]})

        # ── ② 지수 간 방향 불일치 — 대형·기술·소형이 서로 다른 쪽을 본다.
        ads = {"S&P500": at("ad_sp500", d), "NASDAQ": at("ad_nasdaq", d),
               "러셀2000": at("ad_r2000", d)}
        have = {k: v for k, v in ads.items() if v is not None}
        if len(have) == 3:
            up = [k for k, v in have.items() if v > 1]
            dn = [k for k, v in have.items() if v < 1]
            if up and dn:
                out.append({
                    "date": date_s, "sev": "정보", "cat": EVENT_CAT,
                    "title": "미국 증시 — 규모대별로 방향이 갈림",
                    "value": " · ".join(f"{k} {v:.2f}배" for k, v in have.items()),
                    "rule": "같은 날 세 지수(대형·기술·소형)의 A/D 비율이 1을 사이에 두고 갈림 — 임계값 없음",
                    "tags": ["breadth", "equity"]})

        # ── ③ 쏠림 만장일치 — 세 지수 모두 큰 종목이 끌었다(= 상승 폭이 좁다).
        # 만장일치는 기준값이 아니라 **일치 여부**라 임의성이 없다.
        sk = {"S&P500": at("skew_sp500", d), "NASDAQ": at("skew_nasdaq", d),
              "러셀2000": at("skew_r2000", d)}
        skv = [v for v in sk.values() if v is not None]
        if len(skv) == 3 and all(v > 0 for v in skv):
            out.append({
                "date": date_s, "sev": "주의", "cat": EVENT_CAT,
                "title": "미국 증시 — 세 지수 모두 대형주가 끌어올림 (좁은 상승)",
                "value": " · ".join(f"{k} {v:+.2f}%p" for k, v in sk.items() if v is not None),
                "rule": "같은 날 세 지수 모두 (시총가중평균 − 단순평균)>0 — 만장일치 조건, 임계값 없음",
                "tags": ["breadth", "equity"]})
        elif len(skv) == 3 and all(v < 0 for v in skv):
            out.append({
                "date": date_s, "sev": "정보", "cat": EVENT_CAT,
                "title": "미국 증시 — 세 지수 모두 중소형이 끌어올림 (넓은 상승)",
                "value": " · ".join(f"{k} {v:+.2f}%p" for k, v in sk.items() if v is not None),
                "rule": "같은 날 세 지수 모두 (시총가중평균 − 단순평균)<0 — 만장일치 조건, 임계값 없음",
                "tags": ["breadth", "equity"]})

    catalog = [{
        "cat": EVENT_CAT,
        "rule": "미국 증시 데일리 리포트의 **같은 날 안 비교**(임계값 없음): "
                "① A/D 방향과 52주 신고가−신저가 방향이 반대 "
                "② 대형·기술·소형 지수의 A/D 방향이 갈림 "
                "③ 세 지수의 쏠림(시총가중−단순)이 모두 같은 부호",
        "sev": "주의/정보",
    }] if out or cols["ad_ratio"] is not None else []
    return out, catalog
