# -*- coding: utf-8 -*-
"""BM(자산군 전략 벤치마크) 파서 + CMA 사전계산 계약 (`pipeline/bm.py`).

실파일의 형태 특성(2행 헤더·YYYYMMDD 문자열 날짜·주말 캐리포워드·0=미개시)은
`tests/synth.py` 의 `write_bm` 이 흉내 낸다 — 벤더/기관 값은 한 톨도 없다.
CMA 게시물의 수학(연환산·공분산 대칭·헤지 재구성 항등식)은 여기서 손계산과
독립 재계산으로 대조한다.
"""

from __future__ import annotations

import json
import shutil

import numpy as np
import pandas as pd

import bm
import process
import synth


def _store(**series):
    """build_cma 가 기대하는 최소 형태의 시리즈 저장소."""
    return {k: {"s": v} for k, v in series.items()}


# --------------------------------------------------------------------------
# 파서 — 라우팅·헤더·결측 규약
# --------------------------------------------------------------------------

def test_bm_routing_is_filename_prefix_case_insensitive(tmp_path, synth_dir):
    """`bm`/`BM` 접두사면 확장자 앞 내용과 무관하게 벤치마크 파서로 간다."""
    shutil.copy(synth_dir / "BM지수_synth_20991231.xlsx",
                tmp_path / "bm_소문자_사본.xlsx")
    report = process.load_data_dir(tmp_path)
    assert [r["kind"] for r in report] == ["benchmark"]
    assert set(process.SERIES) == set(synth.BM_KEYS)


def test_bm_group_row_carries_forward_like_merged_cells(parsed):
    """1행 그룹 라벨은 구간 첫 칸에만 있고 빈 칸은 왼쪽을 계승한다.

    계승이 깨지면 `시가` 그룹의 뒷열들이 `장부가` 로 흡수되거나 그룹이 빈다 —
    같은 자산군 이름(국내채권)이 두 그룹에 있으므로 키 충돌로 관측이 사라진다.
    """
    _, P = parsed
    assert "bm:장부가 국내채권" in P.SERIES
    assert "bm:시가 국내채권" in P.SERIES
    a = P.SERIES["bm:장부가 국내채권"]["s"]
    b = P.SERIES["bm:시가 국내채권"]["s"]
    # 두 키는 서로 다른 시리즈여야 한다 (실파일 실측: 전 구간 상이)
    common = a.index.intersection(b.index)
    assert len(common) > 0 and not a[common].equals(b[common])


def test_bm_zero_prefix_is_missing_not_observation(parsed):
    """값 0 = 미개시 결측. 0 이 관측으로 새면 개시일에 ±100% 급변이 생긴다."""
    _, P = parsed
    late = P.SERIES["bm:시가 국내주식"]["s"]
    assert late.index.min() >= pd.Timestamp(synth.BM_LATE_START)
    assert (late > 0).all()
    full = P.SERIES["bm:장부가 국내채권"]["s"]
    assert full.index.min() == pd.Timestamp(synth.BM_START)


def test_bm_dates_are_yyyymmdd_strings_and_junk_rows_skipped(parsed):
    """A열은 YYYYMMDD **문자열**이고, 날짜가 아닌 행(합계 등)은 조용히 무시된다."""
    _, P = parsed
    cal = pd.date_range(synth.BM_START, synth.END)
    s = P.SERIES["bm:장부가 국내채권"]["s"]
    assert len(s) == len(cal)                 # 잡음 행이 관측으로 새지 않았다
    assert s.index.min() == cal[0] and s.index.max() == cal[-1]


def test_bm_datetime_a_column_is_parsed(tmp_path):
    """A열이 **날짜형 셀**이어도 파싱된다 — 문서 계약("YYYYMMDD 문자열 또는 날짜형").

    재점검 실측: openpyxl 은 날짜 셀을 datetime.datetime 으로 돌려주는데 예전 코드는
    pd.Timestamp(하위클래스)만 검사해 날짜형 파일이 **경고 0건으로 0 관측**이 됐다 —
    엑셀에서 날짜 서식으로 재저장하는 순간 CMA 가 조용히 꺼지는 함정이었다.
    """
    import datetime
    import openpyxl
    p = tmp_path / "bm_datetime.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("S")
    ws.append(["조회일자", "장부가", "", "시가"])
    ws.append(["", "국내채권", "해외채권", "국내주식"])
    for i in range(40):
        d = datetime.datetime(2024, 1, 1) + datetime.timedelta(days=i)
        ws.append([d, 100.0 + i, 200.0 + i, 300.0 + i])
    wb.save(p)
    warns = []
    cols = bm.parse(p, warns.append)
    assert [len(pairs) for _, _, pairs in cols] == [40, 40, 40]
    assert not warns

    # 반대로 날짜를 한 행도 못 읽으면 조용히 넘어가지 않고 소리를 낸다
    p2 = tmp_path / "bm_nodates.xlsx"
    wb2 = openpyxl.Workbook(write_only=True)
    ws2 = wb2.create_sheet("S")
    ws2.append(["조회일자", "장부가"])
    ws2.append(["", "국내채권"])
    ws2.append(["합계", 1.0])
    wb2.save(p2)
    warns2 = []
    bm.parse(p2, warns2.append)
    assert any("날짜 행이 0개" in w for w in warns2)


def test_build_cma_survives_constant_column_in_window():
    """창 안 무변동 열 — corr 은 None 이 되고 빌드는 산다.

    재점검 실측: NaN corr 이 그대로 나가면 격리망 **밖**의 json.dump(allow_nan=False)
    가 JSON 15개 전체를 죽였다. NaN 은 0 으로 지어내지 않고 None(정의 불가)으로 낸다.
    """
    me = pd.date_range("2023-01-31", periods=26, freq="ME")
    flat = pd.Series(100.0, index=me)                       # 전 구간 무변동 → 수익률 전부 0
    move = pd.Series(100.0 * (1.01 ** np.arange(26)) * (1 + 0.02 * np.sin(np.arange(26))), index=me)
    out = bm.build_cma(
        _store(**{"bm:시가 국내주식": move, "bm:시가 국내채권": flat}), lambda m: None)
    assert out["active"] is True
    json.dumps(out, allow_nan=False)                        # 여기서 죽던 경로
    w = out["windows"][-1]
    i = out["cols"].index("시가 국내채권")
    assert w["vol_pct"][i] == 0.0
    assert w["corr"][i][1 - i] is None and w["corr"][i][i] is None


# --------------------------------------------------------------------------
# 월말 수익률 — 부분월 가드
# --------------------------------------------------------------------------

def test_month_end_returns_drops_partial_final_month():
    """마지막 달은 관측이 달력 월말에 도달했을 때만 표본에 들어간다.

    실측 동기: BM 파일은 8/10까지, 달러원은 7/20까지였는데 리샘플 라벨(월말)만
    믿으면 그 부분월 수익률이 월간으로 둔갑해 σ·평균이 왜곡된다.
    """
    idx = pd.date_range("2024-01-01", "2024-03-15")
    s = pd.Series(np.linspace(100.0, 130.0, len(idx)), index=idx)
    r = bm._month_end_returns(s)
    assert list(r.index) == [pd.Timestamp("2024-02-29")]   # 3월(부분월)은 없다

    idx2 = pd.date_range("2024-01-01", "2024-03-31")
    s2 = pd.Series(np.linspace(100.0, 130.0, len(idx2)), index=idx2)
    r2 = bm._month_end_returns(s2)
    assert list(r2.index) == [pd.Timestamp("2024-02-29"), pd.Timestamp("2024-03-31")]


# --------------------------------------------------------------------------
# CMA — 비활성 경로 (체인 안전장치)
# --------------------------------------------------------------------------

def test_build_cma_inactive_without_bm_series():
    """BM 파일이 없으면 active:false + 안내 문구 — 게이트·배포는 계속 살아야 한다."""
    out = bm.build_cma(_store(), lambda m: None)
    assert out["active"] is False
    assert "BM" in out["reason"]


def test_build_cma_inactive_below_min_sample():
    """공통 월간 표본 12개월 미만이면 통계를 내지 않는다 (극소표본 CMA 금지)."""
    me = pd.date_range("2025-01-31", periods=8, freq="ME")
    a = pd.Series(np.linspace(100.0, 108.0, 8), index=me)
    warns = []
    out = bm.build_cma(_store(**{"bm:시가 국내주식": a}), warns.append)
    assert out["active"] is False
    assert "개월" in out["reason"]


# --------------------------------------------------------------------------
# CMA — 수학 (손계산 대조)
# --------------------------------------------------------------------------

def test_build_cma_annualization_is_exact():
    """연환산 규약: 평균 ×12, σ ×√12 — +1%/+2% 교대 수익률로 손계산과 대조."""
    me = pd.date_range("2020-01-31", periods=25, freq="ME")
    lvl = [100.0]
    for i in range(24):
        lvl.append(lvl[-1] * (1.01 if i % 2 == 0 else 1.02))
    a = pd.Series(lvl, index=me)
    b = pd.Series(list(reversed(lvl)), index=me)
    out = bm.build_cma(
        _store(**{"bm:장부가 국내채권": a, "bm:시가 국내채권": b}), lambda m: None)
    assert out["active"] is True
    assert out["fx_col"] is None and out["cols"] == out["labels"]
    assert out["alt"] is None                    # DESMOOTH 대상이 없는 저장소
    # 장부가 국내채권 → 시가 국내채권 (경제적 실질 관점 치환쌍)
    assert out["econ_map"]["장부가 국내채권"] == "시가 국내채권"

    w = out["windows"][-1]
    assert w["key"] == "all" and w["n_months"] == 24
    r = np.array([0.01, 0.02] * 12)
    # 게시값은 % 소수 4자리 반올림(rd(·,4)) — 허용오차는 그 양자화 반칸(5e-5)
    assert abs(w["mean_pct"][0] - r.mean() * 12 * 100) < 5e-5
    assert abs(w["vol_pct"][0] - r.std(ddof=1) * np.sqrt(12) * 100) < 5e-5


def test_build_cma_shape_on_parsed_store(parsed):
    """합성 전체 저장소에서: 열 순서·행렬 크기·대칭·σ↔공분산 정합·창 그리드."""
    _, P = parsed
    out = bm.build_cma(P.SERIES, lambda m: None)
    assert out["active"] is True
    assert out["labels"] == [k[3:] for k in synth.BM_UNIVERSE]
    # 대체투자(DESMOOTH)와 달러원이 있으니 보조축 둘 다, 이 순서로
    assert out["cols"] == out["labels"] + ["_alt", "_fx"]
    assert out["alt"]["label"] == "시가 대체투자" and -0.9 < out["alt"]["alpha"] < 0.9

    keys = [w["key"] for w in out["windows"]]
    assert keys[-1] == "all" and len(keys) == len(set(keys))
    assert set(keys[:-1]) <= {str(y) for y in bm.WINDOW_YEARS}

    n = len(out["cols"])
    for w in out["windows"]:
        assert len(w["mean_pct"]) == n and len(w["vol_pct"]) == n
        assert len(w["corr"]) == n and all(len(row) == n for row in w["corr"])
        assert len(w["cov"]) == n and all(len(row) == n for row in w["cov"])
        for i in range(n):
            assert abs(w["corr"][i][i] - 1.0) < 1e-9
            assert abs(w["cov"][i][i] - (w["vol_pct"][i] / 100.0) ** 2) < 1e-5
            for j in range(n):
                assert abs(w["cov"][i][j] - w["cov"][j][i]) < 1e-12

    # 공통 표본은 늦개시 자산(0 채움) 이후 + 완결월만
    w_all = out["windows"][-1]
    assert w_all["start"] >= synth.BM_LATE_START
    assert w_all["n_months"] >= 12
    for c in out["coverage"]:
        assert set(c) == {"label", "group", "first", "last", "n_months", "included"}


def test_excluded_asset_classes_are_parsed_but_kept_out_of_the_matrix(parsed):
    """금융상품·대출금은 **배분 대상이 아닐 뿐 없는 자산이 아니다** (2026-08-11 지시).

    행렬에서는 빠지되 파싱·coverage 에는 남아야 한다 — 게시물에서 통째로 지우면
    "파일에는 10개인데 화면은 8개"인 이유가 화면에서 사라진다. 남는 8개는 §7.7
    합의문의 자산군 8개와 일치해야 한다.
    """
    _, P = parsed
    for k in synth.BM_KEYS:
        assert k in P.SERIES, f"제외 자산군도 파싱은 돼야 한다 — {k}"

    out = bm.build_cma(P.SERIES, lambda m: None)
    assert len(out["labels"]) == 8
    assert set(out["excluded"]) == set(bm.EXCLUDED)
    for lb in bm.EXCLUDED:
        assert lb not in out["labels"], f"{lb} 가 행렬에 남아 있습니다"
        assert lb not in out["econ_map"]
    cov = {c["label"]: c["included"] for c in out["coverage"]}
    assert set(cov) == {k[3:] for k in synth.BM_KEYS}, "coverage 는 10개 전부여야 한다"
    assert all(cov[lb] is False for lb in bm.EXCLUDED)
    assert all(cov[k[3:]] is True for k in synth.BM_UNIVERSE)


def test_cma_fx_column_supports_hedged_reconstruction(parsed):
    """`_fx` 축의 존재 이유 — 헤지 반영 분산의 폐형 재구성이 실제로 맞는지.

    같은 공통 표본에서 h=1 완전헤지 수익률(r − r_fx)의 분산을 직접 계산한 값과,
    게시된 공분산 행렬로 재구성한 Σii − 2Σif + Σff 가 일치해야 한다. 이게 깨지면
    `_fx` 열이 자리만 차지하는 오라벨이라는 뜻이다.
    """
    _, P = parsed
    out = bm.build_cma(P.SERIES, lambda m: None)
    w = out["windows"][-1]
    cols = out["cols"]
    i, f = cols.index("시가 해외주식"), cols.index("_fx")

    rets = {"a": bm._month_end_returns(P.SERIES["bm:시가 해외주식"]["s"]),
            "f": bm._month_end_returns(P.SERIES[bm.FX_KEY]["s"])}
    df = pd.DataFrame(rets).dropna()
    df = df.loc[pd.Timestamp(w["start"]):pd.Timestamp(w["end"])]
    assert len(df) == w["n_months"]        # 표본이 다르면 아래 비교는 무의미하다

    direct = float((df["a"] - df["f"]).var(ddof=1) * 12.0)
    rebuilt = w["cov"][i][i] - 2.0 * w["cov"][i][f] + w["cov"][f][f]
    assert abs(direct - rebuilt) < 1e-6


def test_cma_alt_desmoothing_column_matches_hand_calc():
    """`_alt` 보조축 — Geltner AR(1) 역필터를 손계산과 대조.

    관측 수익률 x 를 결정론적으로 만들고, α(=ρ1)와 역필터 y=(x_t−αx_{t−1})/(1−α)
    를 테스트 안에서 numpy 로 독립 계산해 게시된 α·σ(_alt) 와 맞춘다.
    """
    me = pd.date_range("2020-01-31", periods=37, freq="ME")
    x = np.array(([0.01, 0.01, 0.02, 0.02] * 9), dtype=float)   # ρ1 > 0 인 결정 수열
    lvl = 100.0 * np.cumprod(np.r_[1.0, 1.0 + x])
    alt = pd.Series(lvl, index=me)
    other = pd.Series(np.linspace(100.0, 120.0, 37), index=me)
    out = bm.build_cma(
        _store(**{"bm:시가 대체투자": alt, "bm:시가 국내주식": other}), lambda m: None)

    m = x.mean()
    a = float(((x[1:] - m) * (x[:-1] - m)).sum() / ((x - m) ** 2).sum())
    y = (x[1:] - a * x[:-1]) / (1.0 - a)
    assert out["alt"] == {"label": "시가 대체투자", "alpha": round(a, 6), "n_fit": 36}
    assert out["cols"] == out["labels"] + ["_alt"]              # 달러원 없는 저장소
    w = out["windows"][-1]
    i = out["cols"].index("_alt")
    assert w["n_months"] == len(y)                              # 공통 표본 = y 구간
    assert abs(w["vol_pct"][i] - y.std(ddof=1) * np.sqrt(12) * 100) < 5e-5
    assert abs(w["mean_pct"][i] - y.mean() * 12 * 100) < 5e-5


def test_cma_alt_skipped_below_min_sample_without_breaking_cma():
    """DESMOOTH 대상이 24개월 미만이면 보조축만 조용히 빠지고 CMA 는 산다."""
    me = pd.date_range("2024-01-31", periods=20, freq="ME")
    long_me = pd.date_range("2023-01-31", periods=32, freq="ME")
    warns = []
    out = bm.build_cma(_store(**{
        "bm:시가 대체투자": pd.Series(np.linspace(100, 105, 20), index=me),
        "bm:시가 국내주식": pd.Series(np.linspace(100, 130, 32), index=long_me),
    }), warns.append)
    assert out["active"] is True
    assert out["alt"] is None and "_alt" not in out["cols"]
    assert any("디스무딩 보조축 생략" in w for w in warns)


def test_cma_window_mdd_matches_hand_calc():
    """자산별 실측 MDD — 알려진 수익률 경로로 손계산과 대조.

    a: 0×5 → +10% → −20% → +5% → 0×5 — 누적 1.1 → 0.88, MDD = 1 − 0.88/1.1 = 20%.
    b: 첫 달부터 −10% — 초기 원금(1.0)이 고점 후보여야 10% 가 잡힌다(누적 고점만
    보면 0 이 된다 — 그 회귀를 여기서 막는다).
    """
    ra = [0.0] * 5 + [0.10, -0.20, 0.05] + [0.0] * 5
    rb = [-0.10] + [0.0] * 12
    me = pd.date_range("2024-01-31", periods=14, freq="ME")
    lvl = lambda rs: pd.Series(100.0 * np.cumprod([1.0] + [1 + r for r in rs]), index=me)
    out = bm.build_cma(
        _store(**{"bm:시가 국내주식": lvl(ra), "bm:시가 국내채권": lvl(rb)}), lambda m: None)
    assert out["active"] is True
    w = out["windows"][-1]
    ia, ib = out["cols"].index("시가 국내주식"), out["cols"].index("시가 국내채권")
    assert abs(w["mdd_pct"][ia] - (1 - 0.88 / 1.1) * 100) < 5e-4
    assert abs(w["mdd_pct"][ib] - 10.0) < 5e-4


def test_cma_tv_rolling_covariances_consistent_with_fixed_windows(parsed):
    """시변(롤링) 블록 — 개수·모양이 맞고, 마지막 롤링 창 = 같은 길이의 고정 창.

    마지막 대조가 핵심이다: 롤링 경로의 종점과 고정 창은 **같은 표본**이므로 반올림
    까지 정확히 같아야 한다 — 다르면 둘 중 하나의 표본 절단이 틀린 것이다.
    """
    _, P = parsed
    out = bm.build_cma(P.SERIES, lambda m: None)
    n_all = out["windows"][-1]["n_months"]
    n = len(out["cols"])
    assert out["tv"], "롤링 블록이 비어 있다"
    for blk in out["tv"]:
        need = blk["n_window"]
        assert need == int(blk["key"]) * 12
        assert len(blk["dates"]) == len(blk["cov"]) == n_all - need + 1
        assert blk["dates"] == sorted(blk["dates"])          # 시간 오름차순
        for M in (blk["cov"][0], blk["cov"][-1]):
            assert len(M) == n and all(len(r) == n for r in M)
            for i in range(n):
                assert M[i][i] >= 0
                for j in range(n):
                    assert abs(M[i][j] - M[j][i]) < 1e-12
        wfix = next((w for w in out["windows"] if w["key"] == blk["key"]), None)
        assert wfix is not None
        assert blk["cov"][-1] == wfix["cov"], f"{blk['key']}년: 롤링 종점 ≠ 고정 창"
        assert blk["dates"][-1] == wfix["end"]


# --------------------------------------------------------------------------
# CMA — 유출 가드 (원본 수익률·수준 미게시)
# --------------------------------------------------------------------------

def test_cma_block_publishes_no_timeseries(parsed):
    """cma 블록의 수치 배열은 열 수(K+1)를 넘을 수 없다 — 넘으면 시계열 유출이다.

    `test_contract.py` 의 alloc 유출 가드(상한 25)보다 촘촘한 전용 상한이다.
    """
    _, P = parsed
    out = bm.build_cma(P.SERIES, lambda m: None)
    cap = len(out["cols"])

    def walk(o, path="cma"):
        if isinstance(o, list):
            if o and all(isinstance(x, (int, float)) for x in o):
                assert len(o) <= cap, f"{path}: 수치 배열 길이 {len(o)} — 유출 의심"
            for k, v in enumerate(o):
                walk(v, f"{path}[{k}]")
        elif isinstance(o, dict):
            assert "t" not in o and "v" not in o, f"{path}: 시계열 페이로드 금지"
            for k, v in o.items():
                walk(v, f"{path}.{k}")

    walk(out)
    json.dumps(out)    # 직렬화 가능 = 넘파이 스칼라 누수 없음
