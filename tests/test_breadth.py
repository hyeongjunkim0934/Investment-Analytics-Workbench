# -*- coding: utf-8 -*-
"""미국 증시 데일리 리포트 파서 — 집계만 읽고 종목 단위는 읽지 않는가.

이 파서는 다른 파서와 위험의 성격이 다르다. 입력 파일에는 티커·회사명·현재가·
시가총액이 7,000종목 규모로 들어 있고, 이 저장소의 산출물은 **공개 페이지로 나간다.**
그래서 "값을 맞게 읽는가"만큼 **"안 읽어야 할 것을 안 읽는가"** 가 계약이다.

합성 픽스처만 쓴다 — 벤더 값은 한 톨도 들어가지 않는다.
"""

from __future__ import annotations

import openpyxl
import pytest

import breadth
import synth


@pytest.fixture
def report(tmp_path):
    p = tmp_path / "Daily_Stock_Report_synth.xlsx"
    synth.write_stock_report(p)
    return p


def _parse(path, warns=None):
    return breadth.parse(path, lambda m: (warns if warns is not None else []).append(m))


# ---- 판정 --------------------------------------------------------------------
def test_report_is_recognised_by_sheet_not_filename(tmp_path):
    """판정은 **내용**으로 한다 — 리포트는 날짜가 박힌 이름으로 배포된다.

    파일명 접두사 규약을 매일 지키게 하면 언젠가 반드시 잊고, 잊은 날의 파일은
    조용히 지수 익스포트 파서로 흘러가 무의미한 키를 만들거나 그냥 버려진다.
    """
    weird = tmp_path / "완전히-다른-이름-2030.xlsx"
    synth.write_stock_report(weird)
    assert breadth.is_stock_report(weird) is True

    other = tmp_path / "Daily_Stock_Report_아님.xlsx"      # 이름만 그럴듯한 파일
    synth.write_index(other)
    assert breadth.is_stock_report(other) is False


def test_is_stock_report_survives_a_broken_file(tmp_path):
    """열리지 않는 파일에 대해 예외를 던지지 않는다 — 파이프라인 전체가 멈추면 안 된다."""
    bad = tmp_path / "깨진파일.xlsx"
    bad.write_bytes(b"not a zip")
    assert breadth.is_stock_report(bad) is False


# ---- 값 ----------------------------------------------------------------------
def test_observation_date_comes_from_the_body_not_the_filename(tmp_path):
    """관측일은 본문이 밝힌 날짜다.

    실제 파일명은 `Daily_Stock_Report_26.08.04.xlsx`(작성일)인데 본문은
    "미국 2026-08-03 종가 기준"이다. 파일명을 믿으면 **모든 관측이 하루씩 밀린다** —
    시계열로 쌓는 순간 다른 시리즈와의 정렬이 전부 어긋나는 종류의 오류다.
    """
    p = tmp_path / "Daily_Stock_Report_99.12.31.xlsx"
    synth.write_stock_report(p, date="2030-05-17")
    assert _parse(p)["date"] == "2030-05-17"


def test_market_wide_metrics_are_computed_not_copied(report):
    """게시하는 것은 벤더 셀이 아니라 **계산된 비율·차이**다."""
    v = _parse(report)["values"]
    assert v["ad_ratio"] == pytest.approx(3000 / 1000, abs=1e-9)
    assert v["net_new_high"] == 200 - 500
    assert v["net_new_high_pct"] == pytest.approx((200 - 500) / 4000 * 100, abs=1e-9)
    assert v["surge_net"] == 700 - 100
    assert v["participation"] == 70.0
    assert v["universe"] == 4000


def test_numbers_survive_vendor_formatting(tmp_path):
    """`"4,596"` · `"74%"` 같은 문자열 셀을 수로 읽는다(리포트는 서식 문자열로 준다)."""
    p = tmp_path / "r.xlsx"
    synth.write_stock_report(p, market={"total": 12345, "adv": 2000, "dec": 500, "part": 88})
    v = _parse(p)["values"]
    assert v["universe"] == 12345 and v["participation"] == 88.0
    assert v["ad_ratio"] == pytest.approx(4.0)


def test_skew_is_weighted_minus_simple_in_percentage_points(report):
    """쏠림 = 시총가중 − 단순평균. **부호가 결론이다** — 양수면 대형주가 끌었다는 뜻.

    리포트는 소수(0.0155)로 주므로 ×100 해서 %p 로 게시한다. 배율을 빠뜨리면
    화면의 「+0.5%p」가 「+0.005%p」로 나가 사실상 0 으로 읽힌다.
    """
    v = _parse(report)["values"]
    assert v["skew_sp500"] == pytest.approx((0.015 - 0.010) * 100, abs=1e-6)
    assert v["skew_nasdaq"] == pytest.approx((0.018 - 0.020) * 100, abs=1e-6)
    assert v["skew_sp500"] > 0 > v["skew_nasdaq"], "부호가 뒤집혔다"


def test_only_the_three_declared_indexes_are_published(report):
    """지수는 선언된 셋(대형·기술·소형)만 — DOW 는 픽스처에 있어도 나가면 안 된다."""
    v = _parse(report)["values"]
    assert {"ad_sp500", "ad_nasdaq", "ad_r2000"} <= set(v)
    assert not [k for k in v if k.endswith("_dow")], sorted(v)


def test_blocks_are_found_by_header_not_by_row_number(tmp_path):
    """벤더가 줄을 끼워 넣어도 같은 값을 읽어야 한다.

    행을 박아 두면 서식이 한 줄만 밀렸을 때 **조용히 다른 블록을 읽는다** —
    값은 나오는데 뜻이 달라지는 최악의 실패다.
    """
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    synth.write_stock_report(a, header_shift=0)
    synth.write_stock_report(b, header_shift=7)
    assert _parse(a)["values"] == _parse(b)["values"]


# ---- 공개 범위 (이 파일의 핵심) ------------------------------------------------
def test_parser_never_reads_the_per_ticker_sheets(tmp_path):
    """종목 단위 시트가 **있어도** 뽑아내지 않는다.

    이것이 이 파서의 가장 중요한 계약이다. 실제 리포트에는 `S&P500 Daily` 등
    상세 시트에 티커·회사명·현재가·시가총액이 들어 있고, 산출물은 공개 페이지로
    나간다. 픽스처에 일부러 그런 시트를 넣고, 그 안의 어떤 문자열도 결과에
    나타나지 않는지 확인한다.
    """
    p = tmp_path / "r.xlsx"
    synth.write_stock_report(p)
    # 상세 시트를 덧붙인다 — 값은 전부 가짜지만 '종목 단위'라는 형태가 같다.
    wb = openpyxl.load_workbook(p)
    ws = wb.create_sheet("S&P500 Daily")
    ws.append(["S&P500 상세"])
    ws.append(["티커", "회사명", "섹터/산업", "현재가", "변동률", "시가총액"])
    for tk in ("ZZTOP", "QQFAKE", "WWNULL"):
        ws.append([tk, f"{tk} 주식회사", "합성섹터", 123.45, 0.0678, 9_876_543])
    wb.save(p)

    got = _parse(p)
    blob = repr(got)
    for leaked in ("ZZTOP", "QQFAKE", "WWNULL", "주식회사", "123.45", "9876543"):
        assert leaked not in blob, f"종목 단위 값이 새어 나왔다: {leaked}"
    # 그래도 집계는 정상적으로 읽혀야 한다(상세 시트가 파싱을 방해하지 않는다)
    assert got["values"]["ad_ratio"] == pytest.approx(3.0)


def test_parser_opens_only_the_aggregate_sheet(report, monkeypatch):
    """상세 시트를 **열지도 않는다** — 결과에 안 나오는 것보다 강한 계약이다.

    앞 테스트는 산출물에 종목 문자열이 없는지만 본다. 그것만으로는 파서가 상세
    시트를 전부 훑고 나서 우연히 아무것도 못 건진 경우를 구분하지 못한다
    (뮤테이션으로 실제로 통과했다). 여기서는 워크북을 감싸서 **어느 시트가
    읽혔는지**를 기록한다 — 공개 저장소에서 이 계약은 값 정확도만큼 중요하다.
    """
    touched: list[str] = []

    class _Sheet:
        def __init__(self, ws, title):
            self._ws, self._title = ws, title

        def iter_rows(self, *a, **kw):
            touched.append(self._title)
            return self._ws.iter_rows(*a, **kw)

    class _Book:
        def __init__(self, wb):
            self._wb = wb
            self.sheetnames = wb.sheetnames

        def __getitem__(self, name):
            return _Sheet(self._wb[name], name)

        @property
        def worksheets(self):
            return [_Sheet(ws, ws.title) for ws in self._wb.worksheets]

        def close(self):
            self._wb.close()

    real = openpyxl.load_workbook
    monkeypatch.setattr(breadth.openpyxl, "load_workbook",
                        lambda *a, **kw: _Book(real(*a, **kw)))
    # 상세 시트를 붙여 둔다 — 있어도 안 읽어야 한다
    wb = real(report)
    ws = wb.create_sheet("S&P500 Daily")
    ws.append(["티커", "회사명", "현재가"])
    ws.append(["ZZTOP", "합성 주식회사", 123.45])
    wb.save(report)

    assert _parse(report)["values"], "테스트 전제가 깨졌다 — 파싱 자체가 실패했다"
    assert set(touched) == {breadth.SIGNATURE_SHEET}, (
        f"집계 시트 외의 시트를 읽었다: {sorted(set(touched))}"
    )


def test_published_metric_names_are_all_declared(report):
    """뽑은 키가 전부 `METRICS` 에 선언돼 있어야 한다.

    선언되지 않은 키가 나가면 카탈로그에 이름·단위 없이 실리고, 화면은 무슨 수인지
    말하지 못한다 — 「모든 숫자는 그 자리에서 한 줄로 설명한다」 규약 위반이다.
    """
    v = _parse(report)["values"]
    undeclared = sorted(set(v) - set(breadth.METRICS))
    assert not undeclared, f"METRICS 에 없는 키: {undeclared}"


# ---- 실패 처리 ----------------------------------------------------------------
def test_missing_blocks_warn_and_do_not_crash(tmp_path):
    """블록이 없으면 경고를 남기고 빈 결과를 준다 — 조용히 삼키지 않는다."""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = breadth.SIGNATURE_SHEET
    wb.active["A1"] = "미국 2030-01-02 종가 기준"
    wb.save(p)
    warns = []
    assert _parse(p, warns) == {}
    assert warns, "블록이 없는데 경고가 없다"


def test_a_report_without_a_date_is_rejected(tmp_path):
    """관측일이 없으면 값을 쓰지 않는다 — 날짜 없는 관측은 시계열에 넣을 수 없다."""
    p = tmp_path / "nodate.xlsx"
    synth.write_stock_report(p)
    wb = openpyxl.load_workbook(p)
    ws = wb[breadth.SIGNATURE_SHEET]
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=1):
        if row[0].value and "종가 기준" in str(row[0].value):
            row[0].value = "미국 (날짜 미상) 종가"
    wb.save(p)
    warns = []
    assert _parse(p, warns) == {}
    assert any("관측일" in w for w in warns), warns


# ---- 이벤트 규칙 --------------------------------------------------------------
# 위험 요인과 달리 **이력이 필요 없다** — 전부 같은 날 안에서 두 수를 비교하는
# 횡단면 조건이기 때문이다. 그리고 **임의 기준이 하나도 없어야 한다**(상위 제약).

import pandas as pd


def _store(**metrics):
    """`{키: 값}` → `us:` 시리즈 저장소 한 날짜치."""
    ts = pd.Timestamp("2030-05-17")
    return {f"us:{k}": {"s": pd.Series([v], index=[ts])} for k, v in metrics.items()}


WIN = pd.Timestamp("2030-01-01")

#: 세 지수가 전부 같은 방향이고 쏠림도 만장일치가 아닌 '조용한' 배경 —
#: 규칙 하나만 떼어 보려고 나머지를 잠재운다.
QUIET = dict(ad_sp500=2.0, ad_nasdaq=2.0, ad_r2000=2.0,
             skew_sp500=0.1, skew_nasdaq=-0.1, skew_r2000=0.1)


def _fire(**metrics):
    ev, _ = breadth.detect_events(_store(**metrics), WIN)
    return ev


def test_divergence_event_fires_on_opposite_directions():
    """오른 종목이 많은데 52주 신저가가 더 많으면 — 지수와 시장 내부가 다른 말을 한다."""
    ev = _fire(ad_ratio=2.86, net_new_high=-366, universe=4596, **QUIET)
    d = [e for e in ev if "신저가가 더 많음" in e["title"]]
    assert len(d) == 1, [e["title"] for e in ev]
    assert d[0]["sev"] == "주의" and d[0]["cat"] == breadth.EVENT_CAT
    assert "2.86" in d[0]["value"] and "-366" in d[0]["value"]


def test_divergence_event_has_no_threshold():
    """**1종목 차이로도 발동한다** — 이것이 임의 기준이 없다는 증거다.

    「N개 이상이면」 류의 수를 지금 정하면 그 N 은 근거 없는 수가 된다(자의성 금지).
    부호 비교이므로 −1 이든 −366 이든 같은 규칙이 걸린다.
    """
    assert _fire(ad_ratio=1.01, net_new_high=-1, **QUIET), "1종목 차이에서 안 걸린다"
    assert not _fire(ad_ratio=1.01, net_new_high=0, **QUIET), "차이가 0인데 걸렸다"


def test_mirror_case_is_information_not_warning():
    """내린 날인데 신고가가 더 많으면 — 같은 괴리지만 방향이 반대라 '정보'다."""
    ev = _fire(ad_ratio=0.7, net_new_high=+120, **QUIET)
    d = [e for e in ev if "신고가가 더 많음" in e["title"]]
    assert len(d) == 1 and d[0]["sev"] == "정보"


def test_no_divergence_event_when_both_agree():
    """방향이 같으면 아무 일도 아니다 — 규칙이 늘 발동하면 신호가 아니다."""
    assert not [e for e in _fire(ad_ratio=2.0, net_new_high=+200, **QUIET)
                if "신저가가 더 많음" in e["title"] or "신고가가 더 많음" in e["title"]]
    assert not [e for e in _fire(ad_ratio=0.5, net_new_high=-200, **QUIET)
                if "신저가가 더 많음" in e["title"] or "신고가가 더 많음" in e["title"]]


def test_index_disagreement_fires_only_when_split():
    """대형·기술·소형이 1을 사이에 두고 갈릴 때만."""
    base = dict(ad_ratio=1.5, net_new_high=100,
                skew_sp500=0.1, skew_nasdaq=-0.1, skew_r2000=0.1)
    split = _fire(ad_sp500=1.4, ad_nasdaq=0.8, ad_r2000=1.2, **base)
    assert [e for e in split if "방향이 갈림" in e["title"]]
    same = _fire(ad_sp500=1.4, ad_nasdaq=1.8, ad_r2000=1.2, **base)
    assert not [e for e in same if "방향이 갈림" in e["title"]]


def test_skew_event_needs_unanimity_in_both_directions():
    """쏠림은 **만장일치**일 때만 — 만장일치는 기준값이 아니라 일치 여부라 임의성이 없다."""
    base = dict(ad_ratio=1.5, net_new_high=100,
                ad_sp500=2.0, ad_nasdaq=2.0, ad_r2000=2.0)
    up = _fire(skew_sp500=0.5, skew_nasdaq=0.2, skew_r2000=0.01, **base)
    assert [e for e in up if "대형주가 끌어올림" in e["title"]][0]["sev"] == "주의"
    dn = _fire(skew_sp500=-0.5, skew_nasdaq=-0.2, skew_r2000=-0.01, **base)
    assert [e for e in dn if "중소형이 끌어올림" in e["title"]][0]["sev"] == "정보"
    mixed = _fire(skew_sp500=0.5, skew_nasdaq=-0.2, skew_r2000=0.1, **base)
    assert not [e for e in mixed if "끌어올림" in e["title"]], "만장일치가 아닌데 걸렸다"


def test_events_respect_the_lookback_window():
    """창 밖 관측은 이벤트가 되지 않는다 — 다른 규칙과 같은 창을 쓴다."""
    store = _store(ad_ratio=2.0, net_new_high=-300, **QUIET)
    assert breadth.detect_events(store, pd.Timestamp("2030-01-01"))[0]
    assert not breadth.detect_events(store, pd.Timestamp("2031-01-01"))[0]


def test_every_event_states_its_rule_and_says_it_has_no_threshold():
    """규칙 문구가 화면에 그대로 나간다 — 「임계값 없음」은 방어 가능성의 핵심이다."""
    ev = _fire(ad_ratio=2.0, net_new_high=-300,
               ad_sp500=1.4, ad_nasdaq=0.8, ad_r2000=1.2,
               skew_sp500=0.5, skew_nasdaq=0.2, skew_r2000=0.1)
    assert len(ev) == 3, [e["title"] for e in ev]
    for e in ev:
        assert set(e) == {"date", "sev", "cat", "title", "value", "rule", "tags"}, sorted(e)
        assert e["sev"] in ("경계", "주의", "정보")
        assert "임계값 없음" in e["rule"] or "만장일치" in e["rule"], e["rule"]
        assert e["date"] == "2030-05-17"


def test_event_catalog_row_is_published_with_the_rules():
    """규칙 카탈로그에도 실린다 — 화면이 「전 규칙 공개」를 표방하기 때문이다."""
    _, cat = breadth.detect_events(_store(ad_ratio=2.0, net_new_high=-300, **QUIET), WIN)
    assert len(cat) == 1 and cat[0]["cat"] == breadth.EVENT_CAT
    assert "임계값 없음" in cat[0]["rule"]


def test_no_breadth_series_means_no_events():
    """리포트를 안 올렸으면 조용해야 한다 — 없는 데이터로 이벤트를 만들지 않는다."""
    assert breadth.detect_events({}, WIN) == ([], [])
