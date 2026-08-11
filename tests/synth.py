# -*- coding: utf-8 -*-
"""합성(synthetic) 엑셀 픽스처 생성기.

**여기에는 벤더 데이터가 한 톨도 없다.** 모든 값은 고정 시드(`SEED`)의 난수
행보(random walk)로 이 파일 안에서 만들어진다 — 이 저장소는 공개 저장소이므로
비공개 data 저장소의 값이나 그 파생물을 픽스처로 커밋하면 안 된다.
현실감이 필요한 최소한(수준·부호·자릿수)만 흉내 내고, 통계적 성질은 흉내 내지
않는다. 테스트는 "파이프라인이 이 입력을 계약대로 처리하는가"만 본다.

생성하는 워크북 3종 = data 저장소의 3가지 파서 경로:
    data_bb_synth.xlsx    -> parse_wide(p, "bb")      -> 키 `bb:<Notation>`
    data_info_synth.xlsx  -> parse_wide(p, "info")    -> 키 `info:<Notation>`
    ACWI_synth.xlsx       -> parse_index_export(p)    -> 키 `idx:<A1 첫 토큰>`
"""

from __future__ import annotations

import numpy as np
import openpyxl
import pandas as pd

SEED = 20260731
# 위험 지표는 확장창 백분위(min_periods=1000)와 walk-forward 학습(156주 + 5주 엠바고)을
# 쓰므로, risk/hedge 가 실제로 산출물을 내려면 최소 20년 가까운 일별 이력이 필요하다.
START = "2003-01-01"
END = "2026-07-24"


def bdays() -> pd.DatetimeIndex:
    return pd.bdate_range(START, END)


def _walk(rng, n, start, vol, floor=None, mean_rev=None):
    """고정 시드 난수 행보. mean_rev 가 주어지면 그 수준으로 약하게 회귀."""
    x = np.empty(n, dtype=float)
    x[0] = start
    steps = rng.normal(0.0, vol, n)
    for i in range(1, n):
        prev = x[i - 1]
        nxt = prev + steps[i]
        if mean_rev is not None:
            nxt += 0.01 * (mean_rev - prev)
        if floor is not None and nxt < floor:
            nxt = floor + abs(nxt - floor)
        x[i] = nxt
    return x


def _geo(rng, n, start, vol, drift=0.0002):
    r = rng.normal(drift, vol, n)
    return start * np.exp(np.cumsum(r))


# (Notation, category, kind, level) — kind 는 생성 방식만 정한다
BB_SPEC = [
    ("한국_KOSPI_TR", "주식", "geo", 1500.0),
    ("미국_S&P500_TR", "주식", "geo", 1000.0),
    ("미국종합", "채권", "geo", 100.0),
    ("달러원", "환율", "walk", 1200.0),
    ("달러캐나다달러", "환율", "walk", 1.30),
    ("파운드달러", "환율", "walk", 1.30),
    ("미국_하이일드_스프레드", "크레딧", "walk", 4.5),
    ("미국_투자등급_스프레드", "크레딧", "walk", 1.3),
    ("한국_CDS_5년물", "신용", "walk", 45.0),
    ("일본_CDS_5년물", "신용", "walk", 30.0),
    ("독일_CDS_5년물", "신용", "walk", 20.0),
    ("미국_실업률", "거시", "walk", 4.8),
    # 자산배분(alloc.py) 원천 — 국내채권 합성 TR·대체투자 기저에 쓰인다
    ("한국_5y", "금리", "walk", 3.0),
    ("한국_core_CPI_yoy", "물가", "walk", 2.5),
    ("미국_3m", "금리", "walk", 2.0),
    ("유로_3m", "금리", "walk", 1.0),
    ("일본_3m", "금리", "walk", 0.1),
    ("호주_3m", "금리", "walk", 2.5),
    ("캐나다_3m", "금리", "walk", 2.0),
    ("영국_3m", "금리", "walk", 2.0),
    ("유로_5y", "금리", "walk", 1.5),
    ("일본_5y", "금리", "walk", 0.3),
    ("호주_5y", "금리", "walk", 3.0),
    ("캐나다_5y", "금리", "walk", 2.5),
    ("영국_5y", "금리", "walk", 2.5),
]

INFO_SPEC = [
    ("VIX", "변동성", "walk", 18.0),
    ("VKOSPI", "변동성", "walk", 20.0),
    ("한국_3y", "금리", "walk", 2.8),
    ("한국_10y", "금리", "walk", 3.2),
    ("한국_3m", "금리", "walk", 2.5),
    ("Card_AA_plus_3y", "크레딧", "walk", 3.4),
    ("UST2y", "금리", "walk", 2.2),
    ("UST10y", "금리", "walk", 3.0),
    ("UST_ALL_YTM", "금리", "walk", 4.0),   # alloc 해외채권 기대수익(YTM) 관측치
    ("EURKRW", "환율", "walk", 1400.0),
    ("KRWJPY", "환율", "walk", 950.0),
    ("USDCNY", "환율", "walk", 6.9),
    ("AUDKRW", "환율", "walk", 850.0),
    # USD 만 실측 HP 커브를 준다 — hedge 의 "실측(HP)" 분기와 "금리차 프록시"
    # 분기를 한 픽스처에서 모두 지나가게 하려는 의도적 배치다.
    ("SMB_USDKRW_3M", "헤지", "walk", 1.0),   # 3개월 스왑베이시스(연율 %) — hedge 백테스트 입력
    ("USDKRW_HP_3M", "헤지", "walk", 1.1),
    ("USDKRW_HP_6M", "헤지", "walk", 1.2),
    ("USDKRW_HP_12M", "헤지", "walk", 1.3),
]

VOL = {"geo": 0.011, "walk": None}
# 중복 Notation 컬럼에 채우는 센티넬 — 결과에 이 값이 있으면 두 번째 컬럼이 채택된 것
DUP_SENTINEL = -987654.0
_DUP_COL = "__dup__"


def _series_for(rng, spec, n):
    out = {}
    for name, _cat, kind, level in spec:
        if kind == "geo":
            out[name] = _geo(rng, n, level, VOL["geo"])
        else:
            vol = max(abs(level) * 0.004, 0.002)
            floor = 0.01 if level > 0 and level < 10 else None
            out[name] = _walk(rng, n, level, vol, floor=floor, mean_rev=level)
    return out


def write_wide(path, spec, *, dup_notation: str | None = None,
               sheet_title: str = "D", extra_vendor_rows: bool = True) -> None:
    """와이드 익스포트(블룸버그/인포맥스 형식) 워크북을 만든다.

    dup_notation 을 주면 그 Notation 을 한 번 더 넣어 `duplicate column … skipped`
    경고 경로를 재현한다. 중복 컬럼에는 센티넬 값(DUP_SENTINEL)을 채우므로,
    파싱 결과에 그 값이 있으면 "첫 컬럼 채택" 규약이 깨진 것이다.
    """
    idx = bdays()
    rng = np.random.default_rng(SEED)
    data = _series_for(rng, spec, len(idx))
    names = [s[0] for s in spec]
    cats = {s[0]: s[1] for s in spec}
    if dup_notation is not None:
        names = names + [dup_notation]
        data = dict(data)
        data[_DUP_COL] = np.full(len(idx), DUP_SENTINEL, dtype=float)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet_title)
    if extra_vendor_rows:
        # 파서가 무시해야 하는 벤더 잡음 행들 (A열이 날짜도 Category/Notation 도 아님)
        ws.append(["Ticker"] + [f"{n}.TICK" for n in names])
        ws.append(["DATES"] + ["PX_LAST"] * len(names))
    ws.append(["Category"] + [cats[n] for n in names])
    ws.append(["Notation"] + names)
    cols = [data[n] for n in names[:-1]] + [data[_DUP_COL]] \
        if dup_notation is not None else [data[n] for n in names]
    for i, d in enumerate(idx):
        ws.append([d.to_pydatetime()] + [float(c[i]) for c in cols])
    wb.save(path)


# A1 은 "ACWI …" 로 시작해야 한다: 파서가 첫 공백 앞 토큰을 키로 쓰고,
# `idx:ACWI` 는 process/risk/hedge/wf_validation 이 하드코딩해 쓰는 이름이다.
def write_index(path, a1: str = "ACWI  synthetic-fixture", *,
                start=None, end=None, level=300.0, seed_offset=0) -> None:
    """지수 익스포트 워크북 (A1=지수명, `일자`/`종가` 표, B열=종가)."""
    idx = pd.bdate_range(start or START, end or END)
    rng = np.random.default_rng(SEED + seed_offset)
    v = _geo(rng, len(idx), level, 0.010)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append([a1])
    ws.append([])
    ws.append(["일자", "종가"])
    for i, d in enumerate(idx):
        ws.append([d.to_pydatetime(), float(v[i])])
    wb.save(path)


def write_stock_report(path, *, date="2030-05-17", market=None, indexes=None,
                       sheet="Market Overview", header_shift=0) -> None:
    """미국 증시 데일리 리포트 워크북 (집계 블록만 — 종목 단위는 만들지 않는다).

    벤더 값은 한 톨도 들어가지 않는다. 실제 리포트에는 티커·회사명·현재가가
    7,000종목 규모로 들어 있지만, **파서가 그것을 읽지 않는다는 것이 계약**이므로
    픽스처에도 넣지 않는다 — 넣으면 "안 읽는다"를 시험하는 것이 아니라
    "없어서 못 읽는다"를 시험하게 된다. 종목 단위 검사는 별도로 한다.

    `header_shift` 는 벤더가 서식을 바꿔 블록이 밀린 상황을 재현한다 — 파서가
    행 번호가 아니라 헤더 라벨로 블록을 찾는지 확인하는 데 쓴다.
    """
    m = {"total": 4000, "adv": 3000, "dec": 1000, "part": 70,
         "nh": 200, "nl": 500, "up5": 700, "dn5": 100}
    m.update(market or {})
    ix = {"S&P500": (2.0, 0.010, 0.015), "NASDAQ": (3.0, 0.020, 0.018),
          "러셀2000": (4.0, 0.022, 0.017), "DOW": (1.5, 0.011, 0.013)}
    ix.update(indexes or {})

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(sheet)
    ws.append([])
    ws.append(["미국 증시 데일리 리포트"])
    ws.append([f"미국 {date} 종가 기준  ·  합성 픽스처"])
    for _ in range(header_shift):
        ws.append(["(벤더가 끼워 넣은 줄)"])
    ws.append([])
    ws.append(["1. 시장 요약  (미국 상장 보통주 전체)"])
    ws.append(["상장 보통주", "상승", "하락", "참여율", "신고가권", "신저가권",
               "급등 +5%", "급락 -5%", "ETF 상승/하락"])
    ws.append([f"{m['total']:,}", f"{m['adv']:,}", f"{m['dec']:,}", f"{m['part']}%",
               str(m["nh"]), str(m["nl"]), str(m["up5"]), str(m["dn5"]), "900/300"])
    ws.append([])
    ws.append(["4. 지수별 시장 폭 (Breadth)"])
    ws.append(["지수", "종목수", "상승", "하락", "보합", "상승/하락",
               "평균(단순)", "평균(시가총액가중)", "신고가권", "신저가권"])
    for name, (ad, simple, weighted) in ix.items():
        ws.append([name, 500, 340, 160, 0, ad, simple, weighted, 20, 5])
    wb.save(path)


# --- 자산군 전략 벤치마크(BM) — 실파일의 **형태 특성**만 흉내 낸다 -------------
# 2행 헤더(1행 그룹은 병합 셀처럼 빈 칸이 왼쪽 값을 계승), A열 YYYYMMDD 문자열,
# 주말까지 캐리포워드된 일별 수준, 미개시 구간 0 채움(늦게 시작하는 자산 1개).
# 값은 전부 고정 시드 난수다 — 실파일의 수치는 한 톨도 들어가지 않는다.
# 실파일은 2014년부터지만 CMA 공통 표본은 늦개시 자산(아래) 이후만 쓴다 —
# 그 앞 이력은 coverage 숫자에만 나타나므로 픽스처는 파싱 비용만 낮게 2018 시작.
BM_START = "2018-01-01"
#: 늦개시 자산(시가 국내주식 상대역)의 개시일 — 이 전 날짜는 전부 0 이다.
BM_LATE_START = "2022-06-01"
BM_GROUPS = ["장부가"] * 5 + ["시가"] * 5
BM_NAMES = ["국내채권", "해외채권", "금융상품", "단기자금", "대출금",
            "국내주식", "해외주식", "국내채권", "해외채권", "대체투자"]
_BM_VOL = {"장부가": 0.0001, "시가": 0.008}   # 장부가 = 거의 직선(상각), 시가 = 시장


def write_bm(path, *, end=None) -> None:
    """BM 워크북 — 2행 헤더 · YYYYMMDD 문자열 날짜 · 캐리포워드 · 0=미개시 · 잡음 행."""
    end = end or END
    b = pd.bdate_range(BM_START, end)
    cal = pd.date_range(BM_START, end)
    rng = np.random.default_rng(SEED + 77)
    cols = []
    for grp, name in zip(BM_GROUPS, BM_NAMES):
        v = _geo(rng, len(b), 1000.0, _BM_VOL[grp])
        s = pd.Series(v, index=b).reindex(cal).ffill()
        if (grp, name) == ("시가", "국내주식"):
            s[s.index < BM_LATE_START] = 0.0     # 미개시 = 0 (실파일 규약)
        cols.append(s)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("BM지수일자별조회")
    g_row, cur = ["조회일자"], None
    for g in BM_GROUPS:
        g_row.append(g if g != cur else "")      # 병합 셀 흉내 — 구간 첫 칸만 라벨
        cur = g
    ws.append(g_row)
    ws.append([""] + BM_NAMES)
    for i, d in enumerate(cal):
        if i == 400:   # 파서가 조용히 무시해야 하는 벤더 잡음 행
            ws.append(["합계(벤더 잡음 행)"] + [""] * len(BM_NAMES))
        ws.append([d.strftime("%Y%m%d")] + [float(c.iloc[i]) for c in cols])
    wb.save(path)


def build_all(dest) -> dict:
    """다섯 워크북을 dest 에 만들고 {kind: path} 를 돌려준다."""
    dest = str(dest).rstrip("/")
    paths = {
        "bb": f"{dest}/data_bb_synth.xlsx",
        "info": f"{dest}/data_info_synth.xlsx",
        "idx": f"{dest}/ACWI_synth.xlsx",
        "report": f"{dest}/Daily_Stock_Report_synth.xlsx",
        # 실배포처럼 날짜 박힌 이름 — 접두사 `bm`(대소문자 무관)만 지키면 된다
        "bm": f"{dest}/BM지수_synth_20991231.xlsx",
    }
    write_wide(paths["bb"], BB_SPEC, dup_notation="달러원")
    write_wide(paths["info"], INFO_SPEC, sheet_title="DailyRate")
    write_index(paths["idx"])
    write_stock_report(paths["report"])
    write_bm(paths["bm"])
    return paths


BB_KEYS = [f"bb:{s[0]}" for s in BB_SPEC]
INFO_KEYS = [f"info:{s[0]}" for s in INFO_SPEC]
IDX_KEYS = ["idx:ACWI"]
BM_KEYS = [f"bm:{g} {n}" for g, n in zip(BM_GROUPS, BM_NAMES)]
#: CMA 행렬에 실제로 들어가는 자산군 — 배분 대상이 아닌 둘(`bm.EXCLUDED`)을 뺀 8개.
#: 파싱은 10개 그대로이므로 BM_KEYS 와 다르다.
BM_UNIVERSE = [k for k in BM_KEYS
               if k[3:] not in ("장부가 금융상품", "장부가 대출금")]
#: 데일리 리포트가 만드는 키 — `write_stock_report` 기본 픽스처 기준.
#: 지수는 선언된 셋만 나가므로 DOW 는 여기 없다(픽스처에는 있다).
BREADTH_KEYS = [
    "us:ad_ratio", "us:participation", "us:net_new_high", "us:net_new_high_pct",
    "us:surge_net", "us:universe",
    "us:ad_sp500", "us:ad_nasdaq", "us:ad_r2000",
    "us:skew_sp500", "us:skew_nasdaq", "us:skew_r2000",
]
ALL_KEYS = BB_KEYS + INFO_KEYS + IDX_KEYS + BREADTH_KEYS + BM_KEYS
